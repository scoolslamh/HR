from __future__ import annotations

import base64
import hashlib
import io
import json
import os
import shutil
import tempfile
import uuid
from datetime import timedelta
from unittest.mock import patch

from django.contrib.auth import get_user_model
from django.core.files.storage import default_storage
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from accounts.models import Permission, Role, RolePermission, UserRole
from audit.models import AuditLog
from organization.models import (
    Department,
    Employee,
    EmployeeIdentity,
    EmployeeImportBatch,
    EmployeeImportError,
    EmployeeImportRow,
    EmployeePrimaryLocation,
    EmploymentAssignment,
    Location,
)
from organization.services import (
    EMPLOYEE_IMPORT_HEADERS,
    ImportApprovalError,
    ImportFileValidationError,
    SecurityConfigurationError,
    approve_employee_import,
    build_employee_import_template,
    encrypt_sensitive_text,
    national_id_digest,
    preview_employee_import,
)
from organization.services.employee_import import (
    DEPARTMENT_HEADER,
    EMPLOYEE_NAME_HEADER,
    LOCATION_HEADER,
    MANAGER_NAME_HEADER,
    MANAGER_NATIONAL_ID_HEADER,
    MOBILE_HEADER,
    NATIONAL_ID_HEADER,
)


TEST_MEDIA_ROOT = tempfile.mkdtemp(prefix="attendance-employee-import-tests-")
TEST_PII_KEY = base64.urlsafe_b64encode(
    hashlib.sha256(b"employee-import-tests-pii-key").digest()
).decode("ascii")
TEST_HMAC_KEY = base64.urlsafe_b64encode(
    hashlib.sha256(b"employee-import-tests-hmac-key").digest()
).decode("ascii")


def tearDownModule() -> None:
    shutil.rmtree(TEST_MEDIA_ROOT, ignore_errors=True)


@override_settings(
    MEDIA_ROOT=TEST_MEDIA_ROOT,
    PII_ENCRYPTION_KEY=TEST_PII_KEY,
    NATIONAL_ID_HMAC_KEY=TEST_HMAC_KEY,
    PII_ENCRYPTION_KEY_VERSION="test-v1",
    EMPLOYEE_IMPORT_MAX_BYTES=5 * 1024 * 1024,
    EMPLOYEE_IMPORT_MAX_ROWS=5000,
)
class EmployeeImportTestCase(TestCase):
    employee_national_id = "1023456789"
    manager_national_id = "1098765432"

    def setUp(self) -> None:
        self.today = timezone.localdate()
        self.user = get_user_model().objects.create_user(
            username="employee-importer",
            password="test-password-only",
            first_name="مستخدم",
            last_name="الاستيراد",
        )
        self.department = Department.objects.create(
            code="DEPT-IT",
            name_ar="تقنية المعلومات",
            unit_type=Department.UnitType.DEPARTMENT,
            valid_from=self.today - timedelta(days=365),
            created_by=self.user,
            updated_by=self.user,
        )
        self.location = Location.objects.create(
            code="LOC-HQ",
            name_ar="المقر الرئيسي",
            location_type=Location.LocationType.HEADQUARTERS,
            department=self.department,
            created_by=self.user,
            updated_by=self.user,
        )

    def employee_row(self, **overrides: str) -> dict[str, str]:
        row = {
            EMPLOYEE_NAME_HEADER: "نورة العتيبي",
            NATIONAL_ID_HEADER: self.employee_national_id,
            MOBILE_HEADER: "0501234567",
            DEPARTMENT_HEADER: self.department.name_ar,
            LOCATION_HEADER: self.location.name_ar,
            MANAGER_NAME_HEADER: "",
            MANAGER_NATIONAL_ID_HEADER: "",
        }
        row.update(overrides)
        return row

    def workbook_upload(
        self,
        rows: list[dict[str, str]],
        *,
        headers: tuple[str, ...] = EMPLOYEE_IMPORT_HEADERS,
        filename: str | None = None,
    ) -> SimpleUploadedFile:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "الموظفون"
        worksheet.sheet_view.rightToLeft = True
        worksheet.append(list(headers))
        for row in rows:
            worksheet.append([row.get(header, "") for header in headers])

        output = io.BytesIO()
        workbook.save(output)
        workbook.close()
        return SimpleUploadedFile(
            filename or f"employees-{uuid.uuid4().hex}.xlsx",
            output.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )

    def create_employee_with_identity(
        self,
        national_id: str,
        *,
        full_name: str = "موظف موجود",
    ) -> Employee:
        employee = Employee.objects.create(
            employee_number=None,
            full_name_ar=full_name,
            employment_status=Employee.EmploymentStatus.ACTIVE,
            created_by=self.user,
            updated_by=self.user,
        )
        encrypted = encrypt_sensitive_text(
            national_id,
            context=f"employee-national-id:{employee.id}",
        )
        EmployeeIdentity.objects.create(
            employee=employee,
            national_id_hash=national_id_digest(national_id),
            national_id_encrypted=encrypted.ciphertext,
            encryption_key_version=encrypted.key_version,
            national_id_last4=national_id[-4:],
            normalized_length=10,
            verified_at=timezone.now(),
            verification_source=EmployeeIdentity.VerificationSource.IMPORT,
            created_by=self.user,
            updated_by=self.user,
        )
        return employee

    def grant_import_permission(self) -> None:
        permission = Permission.objects.get(code="employees.import")
        role = Role.objects.create(
            code="employee-importer",
            name_ar="مستورد بيانات الموظفين",
            created_by=self.user,
            updated_by=self.user,
        )
        RolePermission.objects.create(
            role=role,
            permission=permission,
            granted_by=self.user,
        )
        UserRole.objects.create(
            user=self.user,
            role=role,
            created_by=self.user,
        )


class EmployeeImportPreviewTests(EmployeeImportTestCase):
    def test_generated_template_is_rtl_and_has_exact_supported_headers(self) -> None:
        content = build_employee_import_template()

        workbook = load_workbook(io.BytesIO(content), read_only=False, data_only=False)
        try:
            worksheet = workbook.active
            headers = tuple(cell.value for cell in next(worksheet.iter_rows(max_row=1)))
            self.assertEqual(headers, EMPLOYEE_IMPORT_HEADERS)
            self.assertTrue(worksheet.sheet_view.rightToLeft)
            self.assertEqual(len(workbook.worksheets), 1)
        finally:
            workbook.close()

    def test_valid_template_is_read_and_arabic_digits_are_normalized(self) -> None:
        arabic_digits = "١٠٢٣٤٥٦٧٨٩"
        upload = self.workbook_upload(
            [self.employee_row(**{NATIONAL_ID_HEADER: f"  {arabic_digits}  "})]
        )

        batch = preview_employee_import(upload, uploaded_by=self.user)

        self.assertEqual(batch.status, EmployeeImportBatch.Status.PREVIEW_READY)
        self.assertEqual(batch.total_rows, 1)
        self.assertEqual(batch.new_rows, 1)
        self.assertEqual(batch.error_rows, 0)
        import_row = batch.rows.get()
        self.assertEqual(import_row.national_id_hash, national_id_digest(self.employee_national_id))
        self.assertEqual(import_row.national_id_last4, "6789")
        self.assertEqual(import_row.display_data_json["national_id_masked"], "******6789")

    def test_reordered_columns_are_mapped_by_header_name(self) -> None:
        reordered_headers = (
            LOCATION_HEADER,
            MANAGER_NATIONAL_ID_HEADER,
            NATIONAL_ID_HEADER,
            EMPLOYEE_NAME_HEADER,
            DEPARTMENT_HEADER,
            MANAGER_NAME_HEADER,
            MOBILE_HEADER,
        )

        batch = preview_employee_import(
            self.workbook_upload(
                [self.employee_row()],
                headers=reordered_headers,
            ),
            uploaded_by=self.user,
        )

        self.assertEqual(batch.status, EmployeeImportBatch.Status.PREVIEW_READY)
        import_row = batch.rows.get()
        self.assertEqual(import_row.display_data_json["full_name_ar"], "نورة العتيبي")
        self.assertEqual(import_row.display_data_json["department_name"], self.department.name_ar)
        self.assertEqual(import_row.display_data_json["location_name"], self.location.name_ar)

    def test_missing_required_column_rejects_file_without_persisting_batch(self) -> None:
        headers = tuple(
            header for header in EMPLOYEE_IMPORT_HEADERS if header != DEPARTMENT_HEADER
        )

        with self.assertRaises(ImportFileValidationError) as raised:
            preview_employee_import(
                self.workbook_upload([self.employee_row()], headers=headers),
                uploaded_by=self.user,
            )

        self.assertEqual(raised.exception.code, "missing_required_headers")
        self.assertEqual(EmployeeImportBatch.objects.count(), 0)
        self.assertEqual(EmployeeImportRow.objects.count(), 0)

    def test_invalid_national_id_is_a_blocking_masked_row_error(self) -> None:
        invalid_value = "12345"
        batch = preview_employee_import(
            self.workbook_upload(
                [self.employee_row(**{NATIONAL_ID_HEADER: invalid_value})]
            ),
            uploaded_by=self.user,
        )

        self.assertEqual(batch.status, EmployeeImportBatch.Status.HAS_ERRORS)
        self.assertEqual(batch.error_rows, 1)
        import_row = batch.rows.get()
        self.assertEqual(import_row.validation_status, EmployeeImportRow.ValidationStatus.ERROR)
        self.assertEqual(import_row.import_action, EmployeeImportRow.ImportAction.SKIP)
        error = batch.errors.get(error_code="invalid_national_id")
        self.assertEqual(error.severity, EmployeeImportError.Severity.ERROR)
        self.assertNotEqual(error.masked_value, invalid_value)
        self.assertNotIn(invalid_value, error.message_ar)

    def test_duplicate_national_id_inside_workbook_blocks_both_rows(self) -> None:
        rows = [
            self.employee_row(**{EMPLOYEE_NAME_HEADER: "نورة العتيبي"}),
            self.employee_row(**{EMPLOYEE_NAME_HEADER: "سارة القحطاني"}),
        ]

        batch = preview_employee_import(
            self.workbook_upload(rows),
            uploaded_by=self.user,
        )

        self.assertEqual(batch.status, EmployeeImportBatch.Status.HAS_ERRORS)
        self.assertEqual(batch.error_rows, 2)
        self.assertEqual(
            batch.errors.filter(error_code="duplicate_national_id").count(),
            2,
        )
        self.assertFalse(
            batch.rows.exclude(
                validation_status=EmployeeImportRow.ValidationStatus.ERROR
            ).exists()
        )

    def test_existing_employee_is_matched_by_hmac_not_by_name(self) -> None:
        employee = self.create_employee_with_identity(
            self.employee_national_id,
            full_name="اسم مختلف تمامًا",
        )

        batch = preview_employee_import(
            self.workbook_upload([self.employee_row()]),
            uploaded_by=self.user,
        )

        self.assertEqual(batch.new_rows, 0)
        self.assertEqual(batch.update_rows, 1)
        import_row = batch.rows.get()
        self.assertEqual(import_row.matched_employee, employee)
        self.assertEqual(import_row.import_action, EmployeeImportRow.ImportAction.UPDATE)

    def test_preview_does_not_modify_employee_or_history(self) -> None:
        employee = self.create_employee_with_identity(
            self.employee_national_id,
            full_name="الاسم قبل المعاينة",
        )
        assignment = EmploymentAssignment.objects.create(
            employee=employee,
            department=self.department,
            valid_from=self.today - timedelta(days=30),
            is_primary=True,
            created_by=self.user,
            updated_by=self.user,
        )
        primary_location = EmployeePrimaryLocation.objects.create(
            employee=employee,
            location=self.location,
            valid_from=self.today - timedelta(days=30),
            created_by=self.user,
            updated_by=self.user,
        )

        preview_employee_import(
            self.workbook_upload(
                [self.employee_row(**{EMPLOYEE_NAME_HEADER: "اسم جديد في الملف"})]
            ),
            uploaded_by=self.user,
        )

        employee.refresh_from_db()
        assignment.refresh_from_db()
        primary_location.refresh_from_db()
        self.assertEqual(employee.full_name_ar, "الاسم قبل المعاينة")
        self.assertIsNone(assignment.valid_to)
        self.assertIsNone(primary_location.valid_to)
        self.assertEqual(Employee.objects.count(), 1)
        self.assertEqual(EmploymentAssignment.objects.count(), 1)
        self.assertEqual(EmployeePrimaryLocation.objects.count(), 1)
        self.assertEqual(AuditLog.objects.count(), 0)


class EmployeeImportApprovalTests(EmployeeImportTestCase):
    def test_employee_without_manager_is_imported_with_department(self) -> None:
        batch = preview_employee_import(
            self.workbook_upload([self.employee_row()]),
            uploaded_by=self.user,
        )

        approved = approve_employee_import(batch, approved_by=self.user)

        self.assertEqual(approved.status, EmployeeImportBatch.Status.APPROVED)
        assignment = EmploymentAssignment.objects.get()
        self.assertEqual(assignment.department, self.department)
        self.assertIsNone(assignment.manager_employee)

    def test_department_is_the_primary_organizational_link(self) -> None:
        batch = preview_employee_import(
            self.workbook_upload([self.employee_row()]),
            uploaded_by=self.user,
        )

        approve_employee_import(batch, approved_by=self.user)

        employee = Employee.objects.get()
        assignment = employee.employment_assignments.get(valid_to__isnull=True)
        self.assertTrue(assignment.is_primary)
        self.assertEqual(assignment.department_id, self.department.id)
        self.assertEqual(
            employee.primary_location_assignments.get(valid_to__isnull=True).location_id,
            self.location.id,
        )

    def test_new_employee_is_created_only_after_approval(self) -> None:
        batch = preview_employee_import(
            self.workbook_upload([self.employee_row()]),
            uploaded_by=self.user,
        )
        self.assertEqual(Employee.objects.count(), 0)

        approved = approve_employee_import(batch, approved_by=self.user)

        self.assertEqual(approved.status, EmployeeImportBatch.Status.APPROVED)
        employee = Employee.objects.get()
        self.assertIsNone(employee.employee_number)
        self.assertEqual(employee.full_name_ar, "نورة العتيبي")
        self.assertEqual(employee.mobile_masked, "05****4567")
        identity = employee.identity
        self.assertEqual(identity.national_id_hash, national_id_digest(self.employee_national_id))
        self.assertEqual(identity.national_id_last4, "6789")
        self.assertEqual(identity.encryption_key_version, "test-v1")
        self.assertNotIn(
            self.employee_national_id.encode("ascii"),
            bytes(identity.national_id_encrypted),
        )
        self.assertEqual(EmploymentAssignment.objects.filter(employee=employee).count(), 1)
        self.assertEqual(EmployeePrimaryLocation.objects.filter(employee=employee).count(), 1)
        self.assertEqual(AuditLog.objects.filter(action="employee_import.approve").count(), 1)

    def test_approval_preserves_department_and_primary_location_history(self) -> None:
        employee = self.create_employee_with_identity(self.employee_national_id)
        old_assignment = EmploymentAssignment.objects.create(
            employee=employee,
            department=self.department,
            valid_from=self.today - timedelta(days=60),
            is_primary=True,
            created_by=self.user,
            updated_by=self.user,
        )
        old_primary_location = EmployeePrimaryLocation.objects.create(
            employee=employee,
            location=self.location,
            valid_from=self.today - timedelta(days=60),
            created_by=self.user,
            updated_by=self.user,
        )
        new_department = Department.objects.create(
            code="DEPT-HR",
            name_ar="الموارد البشرية",
            unit_type=Department.UnitType.DEPARTMENT,
            valid_from=self.today - timedelta(days=365),
            created_by=self.user,
            updated_by=self.user,
        )
        new_location = Location.objects.create(
            code="LOC-BRANCH",
            name_ar="الفرع الشمالي",
            location_type=Location.LocationType.BRANCH,
            department=new_department,
            created_by=self.user,
            updated_by=self.user,
        )
        row = self.employee_row(
            **{
                DEPARTMENT_HEADER: new_department.name_ar,
                LOCATION_HEADER: new_location.name_ar,
            }
        )

        batch = preview_employee_import(
            self.workbook_upload([row]),
            uploaded_by=self.user,
        )
        approve_employee_import(batch, approved_by=self.user)

        old_assignment.refresh_from_db()
        old_primary_location.refresh_from_db()
        self.assertEqual(old_assignment.valid_to, self.today)
        self.assertEqual(old_primary_location.valid_to, self.today)
        current_assignment = employee.employment_assignments.get(valid_to__isnull=True)
        current_location = employee.primary_location_assignments.get(valid_to__isnull=True)
        self.assertEqual(current_assignment.department, new_department)
        self.assertEqual(current_assignment.valid_from, self.today)
        self.assertEqual(current_location.location, new_location)
        self.assertEqual(current_location.valid_from, self.today)
        self.assertEqual(employee.employment_assignments.count(), 2)
        self.assertEqual(employee.primary_location_assignments.count(), 2)

    def test_manager_is_resolved_by_national_id_from_same_workbook(self) -> None:
        manager_row = self.employee_row(
            **{
                EMPLOYEE_NAME_HEADER: "ريم الحربي",
                NATIONAL_ID_HEADER: self.manager_national_id,
                MOBILE_HEADER: "",
            }
        )
        employee_row = self.employee_row(
            **{
                MANAGER_NAME_HEADER: "ريم الحربي",
                MANAGER_NATIONAL_ID_HEADER: self.manager_national_id,
            }
        )

        batch = preview_employee_import(
            self.workbook_upload([employee_row, manager_row]),
            uploaded_by=self.user,
        )
        self.assertEqual(batch.status, EmployeeImportBatch.Status.PREVIEW_READY)
        approve_employee_import(batch, approved_by=self.user)

        employee = EmployeeIdentity.objects.get(
            national_id_hash=national_id_digest(self.employee_national_id)
        ).employee
        manager = EmployeeIdentity.objects.get(
            national_id_hash=national_id_digest(self.manager_national_id)
        ).employee
        self.assertEqual(
            employee.employment_assignments.get(valid_to__isnull=True).manager_employee,
            manager,
        )

    def test_unknown_manager_is_a_warning_and_does_not_block_import(self) -> None:
        unknown_manager_id = "1122334455"
        row = self.employee_row(
            **{
                MANAGER_NAME_HEADER: "مدير غير موجود",
                MANAGER_NATIONAL_ID_HEADER: unknown_manager_id,
            }
        )

        batch = preview_employee_import(
            self.workbook_upload([row]),
            uploaded_by=self.user,
        )

        self.assertEqual(batch.status, EmployeeImportBatch.Status.PREVIEW_READY)
        self.assertEqual(batch.unmatched_manager_rows, 1)
        error = batch.errors.get(error_code="manager_not_found")
        self.assertEqual(error.severity, EmployeeImportError.Severity.WARNING)
        self.assertNotIn(unknown_manager_id, error.message_ar)
        approve_employee_import(batch, approved_by=self.user)
        assignment = EmploymentAssignment.objects.get()
        self.assertEqual(assignment.department, self.department)
        self.assertIsNone(assignment.manager_employee)
        self.assertEqual(Employee.objects.count(), 1)

    def test_invalid_manager_identifier_is_a_warning_and_does_not_block_import(self) -> None:
        row = self.employee_row(
            **{
                MANAGER_NAME_HEADER: "مدير اختياري",
                MANAGER_NATIONAL_ID_HEADER: "12345",
            }
        )

        batch = preview_employee_import(
            self.workbook_upload([row]),
            uploaded_by=self.user,
        )

        self.assertEqual(batch.status, EmployeeImportBatch.Status.PREVIEW_READY)
        warning = batch.errors.get(error_code="invalid_manager_national_id")
        self.assertEqual(warning.severity, EmployeeImportError.Severity.WARNING)
        approve_employee_import(batch, approved_by=self.user)
        assignment = EmploymentAssignment.objects.get()
        self.assertEqual(assignment.department, self.department)
        self.assertIsNone(assignment.manager_employee)

    def test_self_manager_is_a_warning_and_is_not_saved(self) -> None:
        row = self.employee_row(
            **{
                MANAGER_NAME_HEADER: "نورة العتيبي",
                MANAGER_NATIONAL_ID_HEADER: self.employee_national_id,
            }
        )

        batch = preview_employee_import(
            self.workbook_upload([row]),
            uploaded_by=self.user,
        )

        self.assertEqual(batch.status, EmployeeImportBatch.Status.PREVIEW_READY)
        warning = batch.errors.get(error_code="manager_self_reference")
        self.assertEqual(warning.severity, EmployeeImportError.Severity.WARNING)
        approve_employee_import(batch, approved_by=self.user)
        self.assertIsNone(EmploymentAssignment.objects.get().manager_employee)
        self.assertEqual(Employee.objects.count(), 1)

    def test_repeated_approval_is_idempotent(self) -> None:
        batch = preview_employee_import(
            self.workbook_upload([self.employee_row()]),
            uploaded_by=self.user,
        )
        first_result = approve_employee_import(batch, approved_by=self.user)
        counts_after_first = (
            Employee.objects.count(),
            EmployeeIdentity.objects.count(),
            EmploymentAssignment.objects.count(),
            EmployeePrimaryLocation.objects.count(),
            AuditLog.objects.count(),
        )

        second_result = approve_employee_import(batch, approved_by=self.user)

        self.assertEqual(second_result.id, first_result.id)
        self.assertEqual(second_result.status, EmployeeImportBatch.Status.APPROVED)
        self.assertEqual(
            (
                Employee.objects.count(),
                EmployeeIdentity.objects.count(),
                EmploymentAssignment.objects.count(),
                EmployeePrimaryLocation.objects.count(),
                AuditLog.objects.count(),
            ),
            counts_after_first,
        )
        self.assertEqual(AuditLog.objects.filter(action="employee_import.approve").count(), 1)

    def test_approval_rolls_back_every_database_change_on_failure(self) -> None:
        batch = preview_employee_import(
            self.workbook_upload([self.employee_row()]),
            uploaded_by=self.user,
        )

        with patch(
            "organization.services.employee_import._audit_approval",
            side_effect=RuntimeError("forced test failure"),
        ):
            with self.assertRaises(ImportApprovalError) as raised:
                approve_employee_import(batch, approved_by=self.user)

        self.assertEqual(raised.exception.code, "approval_failed")
        batch.refresh_from_db()
        self.assertEqual(batch.status, EmployeeImportBatch.Status.PREVIEW_READY)
        self.assertIsNone(batch.approved_at)
        self.assertIsNone(batch.approved_by)
        self.assertEqual(Employee.objects.count(), 0)
        self.assertEqual(EmployeeIdentity.objects.count(), 0)
        self.assertEqual(EmploymentAssignment.objects.count(), 0)
        self.assertEqual(EmployeePrimaryLocation.objects.count(), 0)
        self.assertEqual(AuditLog.objects.count(), 0)
        self.assertEqual(Department.objects.count(), 1)
        self.assertEqual(Location.objects.count(), 1)


class EmployeeImportDeletionTests(EmployeeImportTestCase):
    def test_delete_requires_authentication_and_import_permission(self) -> None:
        batch = preview_employee_import(
            self.workbook_upload([self.employee_row()]),
            uploaded_by=self.user,
        )
        delete_url = reverse("organization:employee_import_delete", args=(batch.id,))

        anonymous_response = self.client.post(delete_url)
        self.assertEqual(anonymous_response.status_code, 302)
        self.assertIn(reverse("core:login"), anonymous_response.url)
        self.assertTrue(EmployeeImportBatch.objects.filter(id=batch.id).exists())

        self.client.force_login(self.user)
        forbidden_response = self.client.post(delete_url)
        self.assertEqual(forbidden_response.status_code, 403)
        self.assertTrue(EmployeeImportBatch.objects.filter(id=batch.id).exists())

    def test_allowed_batch_deletion_removes_staged_data_file_and_adds_audit(self) -> None:
        batch = preview_employee_import(
            self.workbook_upload(
                [self.employee_row(**{NATIONAL_ID_HEADER: "12345"})]
            ),
            uploaded_by=self.user,
        )
        self.assertEqual(batch.status, EmployeeImportBatch.Status.HAS_ERRORS)
        batch_id = batch.id
        storage_key = batch.storage_key
        self.assertTrue(default_storage.exists(storage_key))
        self.assertTrue(EmployeeImportRow.objects.filter(batch_id=batch_id).exists())
        self.assertTrue(EmployeeImportError.objects.filter(batch_id=batch_id).exists())
        self.grant_import_permission()
        self.client.force_login(self.user)

        detail_response = self.client.get(
            reverse("organization:employee_import_detail", args=(batch_id,))
        )
        self.assertContains(detail_response, "حذف الدفعة")
        self.assertContains(detail_response, "delete-batch-modal")
        delete_url = reverse("organization:employee_import_delete", args=(batch_id,))
        self.assertEqual(self.client.get(delete_url).status_code, 405)

        response = self.client.post(delete_url)

        self.assertRedirects(response, reverse("organization:employee_import_list"))
        self.assertFalse(EmployeeImportBatch.objects.filter(id=batch_id).exists())
        self.assertFalse(EmployeeImportRow.objects.filter(batch_id=batch_id).exists())
        self.assertFalse(EmployeeImportError.objects.filter(batch_id=batch_id).exists())
        self.assertFalse(default_storage.exists(storage_key))
        self.assertEqual(Employee.objects.count(), 0)
        self.assertEqual(Department.objects.count(), 1)
        self.assertEqual(Location.objects.count(), 1)
        audit = AuditLog.objects.get(
            action="employee_import.delete",
            object_id=batch_id,
        )
        self.assertEqual(audit.outcome, AuditLog.Outcome.SUCCESS)
        self.assertEqual(audit.object_repr_masked, "دفعة استيراد موظفين تجريبية")
        self.assertNotIn(self.employee_national_id, json.dumps(audit.before_json))

    def test_approved_batch_deletion_is_blocked_and_keeps_actual_data(self) -> None:
        batch = preview_employee_import(
            self.workbook_upload([self.employee_row()]),
            uploaded_by=self.user,
        )
        approved = approve_employee_import(batch, approved_by=self.user)
        batch_id = approved.id
        storage_key = approved.storage_key
        data_counts = (
            Employee.objects.count(),
            EmployeeIdentity.objects.count(),
            EmploymentAssignment.objects.count(),
            EmployeePrimaryLocation.objects.count(),
            Department.objects.count(),
            Location.objects.count(),
        )
        self.grant_import_permission()
        self.client.force_login(self.user)

        detail_response = self.client.get(
            reverse("organization:employee_import_detail", args=(batch_id,))
        )
        self.assertNotContains(detail_response, "delete-batch-modal")
        response = self.client.post(
            reverse("organization:employee_import_delete", args=(batch_id,)),
            follow=True,
        )

        self.assertContains(response, "لا يمكن حذف دفعة معتمدة")
        self.assertTrue(EmployeeImportBatch.objects.filter(id=batch_id).exists())
        self.assertTrue(default_storage.exists(storage_key))
        self.assertEqual(
            (
                Employee.objects.count(),
                EmployeeIdentity.objects.count(),
                EmploymentAssignment.objects.count(),
                EmployeePrimaryLocation.objects.count(),
                Department.objects.count(),
                Location.objects.count(),
            ),
            data_counts,
        )
        self.assertFalse(
            AuditLog.objects.filter(
                action="employee_import.delete",
                object_id=batch_id,
            ).exists()
        )


class EmployeeImportInterfaceAndPrivacyTests(EmployeeImportTestCase):
    def test_import_pages_require_authentication_and_business_permission(self) -> None:
        url = reverse("organization:employee_import_list")

        anonymous_response = self.client.get(url)
        self.assertEqual(anonymous_response.status_code, 302)
        self.assertIn(reverse("core:login"), anonymous_response.url)

        self.client.force_login(self.user)
        forbidden_response = self.client.get(url)
        self.assertEqual(forbidden_response.status_code, 403)

        self.grant_import_permission()
        allowed_response = self.client.get(url)
        self.assertEqual(allowed_response.status_code, 200)
        self.assertContains(allowed_response, "دفعات الاستيراد")

    def test_full_national_id_is_absent_from_html_audit_and_stored_preview(self) -> None:
        self.grant_import_permission()
        batch = preview_employee_import(
            self.workbook_upload([self.employee_row()]),
            uploaded_by=self.user,
        )
        approve_employee_import(batch, approved_by=self.user)
        self.client.force_login(self.user)

        page_urls = (
            reverse("organization:employee_import_list"),
            reverse("organization:employee_import_detail", args=(batch.id,)),
            reverse("organization:employee_import_preview", args=(batch.id,)),
            reverse("organization:employee_import_errors", args=(batch.id,)),
        )
        rendered_pages = []
        for url in page_urls:
            response = self.client.get(url)
            self.assertEqual(response.status_code, 200)
            rendered = response.content.decode("utf-8")
            rendered_pages.append(rendered)
            self.assertNotIn(self.employee_national_id, rendered)
        self.assertIn("******6789", rendered_pages[2])

        audit_payload = json.dumps(
            list(AuditLog.objects.values()),
            ensure_ascii=False,
            default=str,
        )
        self.assertNotIn(self.employee_national_id, audit_payload)

        import_row = batch.rows.get()
        redacted_preview = json.dumps(
            import_row.display_data_json,
            ensure_ascii=False,
        )
        self.assertNotIn(self.employee_national_id, redacted_preview)
        self.assertNotIn(
            self.employee_national_id.encode("ascii"),
            bytes(import_row.raw_payload_encrypted),
        )

        self.assertTrue(batch.storage_key.endswith(".bin"))
        with default_storage.open(batch.storage_key, "rb") as encrypted_file:
            encrypted_workbook = encrypted_file.read()
        self.assertNotIn(self.employee_national_id.encode("ascii"), encrypted_workbook)

    def test_missing_crypto_configuration_fails_safely_before_persistence(self) -> None:
        upload = self.workbook_upload([self.employee_row()])

        with override_settings(
            PII_ENCRYPTION_KEY="",
            NATIONAL_ID_HMAC_KEY="",
        ):
            with patch.dict(
                os.environ,
                {
                    "PII_ENCRYPTION_KEY": "",
                    "NATIONAL_ID_HMAC_KEY": "",
                },
            ):
                with self.assertRaises(SecurityConfigurationError) as raised:
                    preview_employee_import(upload, uploaded_by=self.user)

        self.assertEqual(raised.exception.code, "security_configuration_missing")
        self.assertNotIn("KEY", raised.exception.message_ar.upper())
        self.assertEqual(EmployeeImportBatch.objects.count(), 0)
        self.assertEqual(Employee.objects.count(), 0)


class EmployeeImportEmployeeManagementTests(EmployeeImportTestCase):
    def setUp(self) -> None:
        super().setUp()
        self.admin = get_user_model().objects.create_superuser(
            username="employee-management-admin",
            password="Strong-Test-Pass-2026",
        )

    def test_admin_adds_employee_manually_from_import_page(self) -> None:
        self.client.force_login(self.admin)

        response = self.client.post(
            reverse("organization:employee_manual_create"),
            {
                "national_id": "1088888888",
                "full_name_ar": "موظف مضاف يدويًا",
                "employee_number": "MANUAL-1",
                "mobile": "0501112233",
                "department": str(self.department.id),
                "location": str(self.location.id),
                "manager_employee": "",
            },
        )

        self.assertRedirects(response, reverse("organization:employee_import_list"))
        employee = Employee.objects.get(employee_number="MANUAL-1")
        self.assertEqual(employee.identity.national_id_last4, "8888")
        self.assertEqual(employee.employment_assignments.get().department, self.department)
        self.assertEqual(employee.primary_location_assignments.get().location, self.location)
        self.assertTrue(
            AuditLog.objects.filter(
                action="employee.create_manual", object_id=employee.id
            ).exists()
        )

    def test_bulk_archive_and_restore_selected_employees(self) -> None:
        self.client.force_login(self.admin)
        self.client.post(
            reverse("organization:employee_manual_create"),
            {
                "national_id": "1077777777",
                "full_name_ar": "موظف للإدارة الجماعية",
                "employee_number": "BULK-1",
                "mobile": "",
                "department": str(self.department.id),
                "location": str(self.location.id),
                "manager_employee": "",
            },
        )
        employee = Employee.objects.get(employee_number="BULK-1")

        archive_response = self.client.post(
            reverse("organization:employee_bulk_manage"),
            {
                "action": "archive",
                "scope": "selected",
                "employee_ids": [str(employee.id)],
                "batch": "",
                "confirmation": "أرشفة المحدد",
                "reason": "أرشفة موظف انتهت حاجته التشغيلية",
            },
        )
        self.assertRedirects(archive_response, reverse("organization:employee_import_list"))
        employee.refresh_from_db()
        self.assertEqual(employee.employment_status, Employee.EmploymentStatus.ARCHIVED)
        self.assertIsNotNone(employee.archived_at)

        restore_response = self.client.post(
            reverse("organization:employee_bulk_manage"),
            {
                "action": "restore",
                "scope": "selected",
                "employee_ids": [str(employee.id)],
                "batch": "",
                "confirmation": "استعادة المحدد",
                "reason": "إعادة الموظف إلى الحالة التشغيلية",
            },
        )
        self.assertRedirects(restore_response, reverse("organization:employee_import_list"))
        employee.refresh_from_db()
        self.assertEqual(employee.employment_status, Employee.EmploymentStatus.ACTIVE)
        self.assertIsNone(employee.archived_at)
        self.assertEqual(
            AuditLog.objects.filter(
                object_id=employee.id,
                action__in=("employee.bulk_archive", "employee.bulk_restore"),
            ).count(),
            2,
        )

    def test_non_admin_cannot_use_employee_management_actions(self) -> None:
        self.grant_import_permission()
        self.client.force_login(self.user)
        list_response = self.client.get(reverse("organization:employee_import_list"))
        self.assertEqual(list_response.status_code, 200)
        self.assertNotContains(list_response, "الإدارة الجماعية للموظفين")
        response = self.client.post(
            reverse("organization:employee_bulk_manage"),
            {
                "action": "archive",
                "scope": "all",
                "confirmation": "أرشفة جميع الموظفين",
                "reason": "محاولة غير مسموحة للإدارة الجماعية",
            },
        )
        self.assertEqual(response.status_code, 403)

    def test_admin_can_archive_employees_in_approved_batch_scope(self) -> None:
        self.client.force_login(self.admin)
        self.client.post(
            reverse("organization:employee_manual_create"),
            {
                "national_id": "1066666666",
                "full_name_ar": "موظف دفعة محددة",
                "employee_number": "BATCH-1",
                "mobile": "",
                "department": str(self.department.id),
                "location": str(self.location.id),
                "manager_employee": "",
            },
        )
        employee = Employee.objects.get(employee_number="BATCH-1")
        batch = EmployeeImportBatch.objects.create(
            original_filename="approved-employees.xlsx",
            storage_key="employee-imports/approved-test.bin",
            file_sha256="b" * 64,
            file_size_bytes=100,
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            encryption_key_version="test-v1",
            status=EmployeeImportBatch.Status.APPROVED,
            approved_by=self.admin,
            approved_at=timezone.now(),
        )
        EmployeeImportRow.objects.create(
            batch=batch,
            row_number=1,
            raw_payload_encrypted=b"encrypted",
            encryption_key_version="test-v1",
            payload_sha256="c" * 64,
            national_id_hash=employee.identity.national_id_hash,
            national_id_last4="6666",
            display_data_json={},
            import_action=EmployeeImportRow.ImportAction.CREATE,
            validation_status=EmployeeImportRow.ValidationStatus.VALID,
            matched_employee=employee,
        )

        response = self.client.post(
            reverse("organization:employee_bulk_manage"),
            {
                "action": "archive",
                "scope": "batch",
                "employee_ids": [],
                "batch": str(batch.id),
                "confirmation": "أرشفة موظفي الدفعة",
                "reason": "أرشفة جميع موظفي الدفعة المحددة",
            },
        )

        self.assertRedirects(response, reverse("organization:employee_import_list"))
        employee.refresh_from_db()
        self.assertEqual(employee.employment_status, Employee.EmploymentStatus.ARCHIVED)
