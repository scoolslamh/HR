from __future__ import annotations

import hashlib
import io
import json
import re
import uuid
from dataclasses import dataclass, field
from datetime import date
from typing import Any

from django.conf import settings
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.db import IntegrityError, transaction
from django.db.models import Q
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from audit.models import AuditLog
from organization.models import (
    Department,
    Employee,
    EmployeeIdentity,
    EmployeeImportBatch,
    EmployeeImportError,
    EmployeeImportRow,
    EmployeePrimaryLocation,
    EmploymentAssignment,
    Location,
)

from .exceptions import (
    DuplicateImportFileError,
    EmployeeImportServiceError,
    ImportApprovalError,
    ImportDeletionError,
    ImportFileValidationError,
    ImportPreviewError,
)
from .identity import (
    decrypt_sensitive_text,
    encrypt_sensitive_bytes,
    encrypt_sensitive_text,
    ensure_crypto_configured,
    mask_mobile,
    mask_national_id,
    mask_untrusted_identifier,
    national_id_digest,
    normalize_national_id,
    normalize_saudi_mobile,
    redact_potential_national_ids,
)
from .xlsx import random_storage_key, validate_xlsx_upload


EMPLOYEE_NAME_HEADER = "اسم الموظف"
NATIONAL_ID_HEADER = "السجل المدني"
MOBILE_HEADER = "رقم الجوال"
DEPARTMENT_HEADER = "القسم"
LOCATION_HEADER = "مكان الحضور والانصراف"
MANAGER_NAME_HEADER = "المدير المباشر"
MANAGER_NATIONAL_ID_HEADER = "السجل المدني للمدير المباشر"

EMPLOYEE_IMPORT_HEADERS = (
    EMPLOYEE_NAME_HEADER,
    NATIONAL_ID_HEADER,
    MOBILE_HEADER,
    DEPARTMENT_HEADER,
    LOCATION_HEADER,
    MANAGER_NAME_HEADER,
    MANAGER_NATIONAL_ID_HEADER,
)
REQUIRED_HEADERS = frozenset(
    {
        EMPLOYEE_NAME_HEADER,
        NATIONAL_ID_HEADER,
        DEPARTMENT_HEADER,
        LOCATION_HEADER,
    }
)

_CONTROL_CHARACTERS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]+")
_SPACE_RUN = re.compile(r"\s+")


@dataclass(slots=True)
class RowIssue:
    error_code: str
    severity: str
    field_name: str
    message_ar: str
    masked_value: str = ""


@dataclass(slots=True)
class PreparedRow:
    row_number: int
    raw_payload: dict[str, str]
    full_name: str
    national_id: str | None
    national_id_hash: str | None
    mobile: str | None
    department_name: str
    location_name: str
    manager_name: str
    manager_national_id: str | None
    manager_national_id_hash: str | None
    matched_employee: Employee | None = None
    issues: list[RowIssue] = field(default_factory=list)

    @property
    def has_errors(self) -> bool:
        return any(
            issue.severity == EmployeeImportError.Severity.ERROR
            for issue in self.issues
        )

    @property
    def has_warnings(self) -> bool:
        return any(
            issue.severity == EmployeeImportError.Severity.WARNING
            for issue in self.issues
        )


def _clean_text(value: object, *, maximum: int | None = None) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        rendered = "نعم" if value else "لا"
    elif isinstance(value, float) and value.is_integer():
        rendered = str(int(value))
    else:
        rendered = str(value)
    rendered = _CONTROL_CHARACTERS.sub("", rendered)
    rendered = _SPACE_RUN.sub(" ", rendered).strip()
    if maximum is not None:
        rendered = rendered[:maximum]
    return rendered


def _normalize_header(value: object) -> str:
    return _clean_text(value)


def _max_rows() -> int:
    value = getattr(settings, "EMPLOYEE_IMPORT_MAX_ROWS", 5000)
    try:
        maximum = int(value)
    except (TypeError, ValueError) as exc:
        raise ImportFileValidationError(
            "حد عدد صفوف الاستيراد غير مضبوط بصورة صحيحة.",
            code="invalid_row_limit",
        ) from exc
    if maximum <= 0:
        raise ImportFileValidationError(
            "حد عدد صفوف الاستيراد غير مضبوط بصورة صحيحة.",
            code="invalid_row_limit",
        )
    return maximum


def _read_rows(content: bytes) -> list[tuple[int, dict[str, str]]]:
    try:
        workbook = load_workbook(
            io.BytesIO(content),
            read_only=True,
            data_only=False,
            keep_links=False,
        )
    except Exception as exc:
        raise ImportFileValidationError(
            "تعذر قراءة ملف xlsx.",
            code="unreadable_workbook",
        ) from exc

    try:
        if len(workbook.worksheets) != 1:
            raise ImportFileValidationError(
                "يجب أن يحتوي قالب الموظفين على ورقة عمل واحدة فقط.",
                code="unexpected_worksheet_count",
            )
        worksheet = workbook.worksheets[0]
        maximum_rows = _max_rows()
        maximum_columns = 50
        if worksheet.max_row > maximum_rows + 1:
            raise ImportFileValidationError(
                "عدد صفوف الملف يتجاوز الحد المسموح.",
                code="too_many_rows",
            )
        if worksheet.max_column > maximum_columns:
            raise ImportFileValidationError(
                "عدد أعمدة الملف يتجاوز الحد المسموح.",
                code="too_many_columns",
            )

        all_rows = worksheet.iter_rows(
            min_row=1,
            max_row=max(worksheet.max_row, 1),
            max_col=max(worksheet.max_column, 1),
        )
        header_cells = next(all_rows, None)
        if header_cells is None:
            raise ImportFileValidationError(
                "لا يحتوي الملف على صف العناوين.",
                code="missing_header_row",
            )
        if any(cell.data_type == "f" for cell in header_cells):
            raise ImportFileValidationError(
                "لا يسمح باستخدام الصيغ في ملف الاستيراد.",
                code="formula_not_allowed",
            )
        headers = [_normalize_header(cell.value) for cell in header_cells]
        populated_headers = [header for header in headers if header]
        if len(populated_headers) != len(set(populated_headers)):
            raise ImportFileValidationError(
                "يحتوي الملف على أسماء أعمدة مكررة.",
                code="duplicate_headers",
            )
        missing_headers = REQUIRED_HEADERS.difference(populated_headers)
        if missing_headers:
            raise ImportFileValidationError(
                "يفتقد الملف واحدًا أو أكثر من الأعمدة الإلزامية.",
                code="missing_required_headers",
            )

        parsed_rows: list[tuple[int, dict[str, str]]] = []
        for row_number, cells in enumerate(all_rows, start=2):
            if any(cell.data_type == "f" for cell in cells):
                raise ImportFileValidationError(
                    "لا يسمح باستخدام الصيغ في ملف الاستيراد.",
                    code="formula_not_allowed",
                )
            values = [_clean_text(cell.value) for cell in cells]
            if not any(values):
                continue
            raw_payload = {
                header: value
                for header, value in zip(headers, values, strict=True)
                if header
            }
            parsed_rows.append((row_number, raw_payload))
        return parsed_rows
    finally:
        workbook.close()


def _required_text(
    raw_payload: dict[str, str],
    header: str,
    field_name: str,
    maximum: int,
    issues: list[RowIssue],
) -> str:
    value = _clean_text(raw_payload.get(header, ""))
    if not value:
        issues.append(
            RowIssue(
                error_code="required_value",
                severity=EmployeeImportError.Severity.ERROR,
                field_name=field_name,
                message_ar=f"حقل {header} مطلوب.",
            )
        )
        return ""
    if len(value) > maximum:
        issues.append(
            RowIssue(
                error_code="value_too_long",
                severity=EmployeeImportError.Severity.ERROR,
                field_name=field_name,
                message_ar=f"قيمة حقل {header} أطول من الحد المسموح.",
            )
        )
    return value[:maximum]


def _prepare_row(row_number: int, raw_payload: dict[str, str]) -> PreparedRow:
    issues: list[RowIssue] = []
    full_name = _required_text(
        raw_payload, EMPLOYEE_NAME_HEADER, "full_name_ar", 250, issues
    )
    department_name = _required_text(
        raw_payload, DEPARTMENT_HEADER, "department", 200, issues
    )
    location_name = _required_text(
        raw_payload, LOCATION_HEADER, "location", 200, issues
    )
    manager_name = _clean_text(raw_payload.get(MANAGER_NAME_HEADER, ""))
    if len(manager_name) > 250:
        issues.append(
            RowIssue(
                error_code="manager_name_too_long",
                severity=EmployeeImportError.Severity.WARNING,
                field_name="manager_name",
                message_ar=(
                    "اسم المدير المباشر أطول من الحد المسموح؛ "
                    "لن يمنع ذلك استيراد الموظف."
                ),
            )
        )
        manager_name = manager_name[:250]

    raw_national_id = raw_payload.get(NATIONAL_ID_HEADER, "")
    national_id: str | None = None
    national_hash: str | None = None
    try:
        national_id = normalize_national_id(raw_national_id)
        national_hash = national_id_digest(national_id)
    except ValueError:
        issues.append(
            RowIssue(
                error_code="invalid_national_id",
                severity=EmployeeImportError.Severity.ERROR,
                field_name="national_id",
                message_ar="يجب أن يتكون السجل المدني من 10 أرقام.",
                masked_value=mask_untrusted_identifier(raw_national_id),
            )
        )

    raw_mobile = raw_payload.get(MOBILE_HEADER, "")
    mobile: str | None = None
    try:
        mobile = normalize_saudi_mobile(raw_mobile)
    except ValueError:
        issues.append(
            RowIssue(
                error_code="invalid_mobile",
                severity=EmployeeImportError.Severity.ERROR,
                field_name="mobile",
                message_ar="رقم الجوال السعودي غير صحيح.",
                masked_value=mask_untrusted_identifier(raw_mobile),
            )
        )

    raw_manager_national_id = raw_payload.get(MANAGER_NATIONAL_ID_HEADER, "")
    manager_national_id: str | None = None
    manager_hash: str | None = None
    if _clean_text(raw_manager_national_id):
        try:
            manager_national_id = normalize_national_id(raw_manager_national_id)
            manager_hash = national_id_digest(manager_national_id)
        except ValueError:
            issues.append(
                RowIssue(
                    error_code="invalid_manager_national_id",
                    severity=EmployeeImportError.Severity.WARNING,
                    field_name="manager_national_id",
                    message_ar=(
                        "السجل المدني للمدير المباشر غير صحيح؛ "
                        "سيُستورد الموظف دون ربط المدير."
                    ),
                    masked_value=mask_untrusted_identifier(raw_manager_national_id),
                )
            )
    elif manager_name:
        issues.append(
            RowIssue(
                error_code="manager_identifier_missing",
                severity=EmployeeImportError.Severity.WARNING,
                field_name="manager_national_id",
                message_ar="لن يُربط المدير لأن سجل المدير المدني غير مدخل.",
            )
        )

    return PreparedRow(
        row_number=row_number,
        raw_payload=raw_payload,
        full_name=full_name,
        national_id=national_id,
        national_id_hash=national_hash,
        mobile=mobile,
        department_name=department_name,
        location_name=location_name,
        manager_name=manager_name,
        manager_national_id=manager_national_id,
        manager_national_id_hash=manager_hash,
        issues=issues,
    )


def _active_departments(name: str, effective_date: date):
    return Department.objects.filter(
        name_ar=name,
        is_active=True,
        archived_at__isnull=True,
        valid_from__lte=effective_date,
    ).filter(Q(valid_to__isnull=True) | Q(valid_to__gt=effective_date))


def _active_locations(name: str):
    return Location.objects.filter(name_ar=name, is_active=True)


def _add_reference_issues(rows: list[PreparedRow], effective_date: date) -> None:
    department_counts = {
        name: _active_departments(name, effective_date).count()
        for name in {row.department_name for row in rows if row.department_name}
    }
    location_counts = {
        name: _active_locations(name).count()
        for name in {row.location_name for row in rows if row.location_name}
    }
    for row in rows:
        department_count = department_counts.get(row.department_name, 0)
        if row.department_name and department_count == 0:
            row.issues.append(
                RowIssue(
                    error_code="missing_department",
                    severity=EmployeeImportError.Severity.WARNING,
                    field_name="department",
                    message_ar="القسم غير موجود، ويلزم إنشاؤه صراحة عند الاعتماد.",
                    masked_value=row.department_name,
                )
            )
        elif department_count > 1:
            row.issues.append(
                RowIssue(
                    error_code="ambiguous_department",
                    severity=EmployeeImportError.Severity.ERROR,
                    field_name="department",
                    message_ar="اسم القسم يطابق أكثر من سجل نشط.",
                    masked_value=row.department_name,
                )
            )

        location_count = location_counts.get(row.location_name, 0)
        if row.location_name and location_count == 0:
            row.issues.append(
                RowIssue(
                    error_code="missing_location",
                    severity=EmployeeImportError.Severity.WARNING,
                    field_name="location",
                    message_ar="الموقع غير موجود، ويلزم إنشاؤه صراحة عند الاعتماد.",
                    masked_value=row.location_name,
                )
            )
        elif location_count > 1:
            row.issues.append(
                RowIssue(
                    error_code="ambiguous_location",
                    severity=EmployeeImportError.Severity.ERROR,
                    field_name="location",
                    message_ar="اسم الموقع يطابق أكثر من سجل نشط.",
                    masked_value=row.location_name,
                )
            )


def _add_identity_and_manager_issues(rows: list[PreparedRow]) -> None:
    hashes = [row.national_id_hash for row in rows if row.national_id_hash]
    duplicate_hashes = {value for value in hashes if hashes.count(value) > 1}
    existing_identities = {
        identity.national_id_hash: identity
        for identity in EmployeeIdentity.objects.select_related("employee").filter(
            national_id_hash__in=set(hashes)
        )
    }
    file_rows_by_hash = {
        row.national_id_hash: row for row in rows if row.national_id_hash
    }
    manager_hashes = {
        row.manager_national_id_hash
        for row in rows
        if row.manager_national_id_hash
    }
    existing_managers = {
        identity.national_id_hash: identity.employee
        for identity in EmployeeIdentity.objects.select_related("employee").filter(
            national_id_hash__in=manager_hashes
        )
    }

    for row in rows:
        if row.national_id_hash in duplicate_hashes:
            row.issues.append(
                RowIssue(
                    error_code="duplicate_national_id",
                    severity=EmployeeImportError.Severity.ERROR,
                    field_name="national_id",
                    message_ar="السجل المدني مكرر داخل الملف.",
                    masked_value=(
                        mask_national_id(row.national_id) if row.national_id else ""
                    ),
                )
            )
        identity = existing_identities.get(row.national_id_hash or "")
        if identity is not None:
            row.matched_employee = identity.employee

        manager_hash = row.manager_national_id_hash
        if not manager_hash:
            continue
        if manager_hash == row.national_id_hash:
            row.issues.append(
                RowIssue(
                    error_code="manager_self_reference",
                    severity=EmployeeImportError.Severity.WARNING,
                    field_name="manager_national_id",
                    message_ar=(
                        "لا يمكن ربط الموظف مديرًا مباشرًا لنفسه؛ "
                        "سيُستورد دون ربط المدير."
                    ),
                    masked_value=(
                        mask_national_id(row.manager_national_id)
                        if row.manager_national_id
                        else ""
                    ),
                )
            )
            continue
        manager = existing_managers.get(manager_hash)
        file_manager_row = file_rows_by_hash.get(manager_hash)
        if manager is None and file_manager_row is None:
            row.issues.append(
                RowIssue(
                    error_code="manager_not_found",
                    severity=EmployeeImportError.Severity.WARNING,
                    field_name="manager_national_id",
                    message_ar=(
                        "تعذرت مطابقة المدير المباشر بواسطة سجله المدني؛ "
                        "سيُستورد الموظف دون ربط المدير."
                    ),
                    masked_value=(
                        mask_national_id(row.manager_national_id)
                        if row.manager_national_id
                        else ""
                    ),
                )
            )
            continue
        expected_name = manager.full_name_ar if manager else file_manager_row.full_name
        if row.manager_name and _clean_text(expected_name) != row.manager_name:
            row.issues.append(
                RowIssue(
                    error_code="manager_name_mismatch",
                    severity=EmployeeImportError.Severity.WARNING,
                    field_name="manager_name",
                    message_ar="اسم المدير للعرض لا يطابق الموظف المحدد بسجله المدني.",
                )
            )


def _display_data(row: PreparedRow) -> dict[str, str]:
    return {
        "full_name_ar": row.full_name,
        "national_id_masked": (
            mask_national_id(row.national_id)
            if row.national_id
            else mask_untrusted_identifier(row.raw_payload.get(NATIONAL_ID_HEADER, ""))
        ),
        "mobile_masked": mask_mobile(row.mobile),
        "department_name": row.department_name,
        "location_name": row.location_name,
        "manager_name": row.manager_name,
        "manager_national_id_masked": (
            mask_national_id(row.manager_national_id)
            if row.manager_national_id
            else ""
        ),
    }


def _save_preview_rows(
    batch: EmployeeImportBatch,
    prepared_rows: list[PreparedRow],
) -> None:
    import_rows: list[EmployeeImportRow] = []
    issues_by_row_id: dict[uuid.UUID, list[RowIssue]] = {}
    for prepared in prepared_rows:
        row_id = uuid.uuid4()
        canonical_payload = json.dumps(
            prepared.raw_payload,
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        )
        encrypted = encrypt_sensitive_text(
            canonical_payload,
            context=f"employee-import-row:{row_id}",
        )
        if prepared.has_errors:
            validation_status = EmployeeImportRow.ValidationStatus.ERROR
            import_action = EmployeeImportRow.ImportAction.SKIP
        elif prepared.has_warnings:
            validation_status = EmployeeImportRow.ValidationStatus.WARNING
            import_action = (
                EmployeeImportRow.ImportAction.UPDATE
                if prepared.matched_employee
                else EmployeeImportRow.ImportAction.CREATE
            )
        else:
            validation_status = EmployeeImportRow.ValidationStatus.VALID
            import_action = (
                EmployeeImportRow.ImportAction.UPDATE
                if prepared.matched_employee
                else EmployeeImportRow.ImportAction.CREATE
            )
        import_rows.append(
            EmployeeImportRow(
                id=row_id,
                batch=batch,
                row_number=prepared.row_number,
                raw_payload_encrypted=encrypted.ciphertext,
                encryption_key_version=encrypted.key_version,
                payload_sha256=hashlib.sha256(
                    canonical_payload.encode("utf-8")
                ).hexdigest(),
                national_id_hash=prepared.national_id_hash,
                national_id_last4=(prepared.national_id[-4:] if prepared.national_id else ""),
                display_data_json=_display_data(prepared),
                import_action=import_action,
                validation_status=validation_status,
                matched_employee=prepared.matched_employee,
            )
        )
        issues_by_row_id[row_id] = prepared.issues
    EmployeeImportRow.objects.bulk_create(import_rows)
    EmployeeImportError.objects.bulk_create(
        [
            EmployeeImportError(
                batch=batch,
                row_id=row.id,
                error_code=issue.error_code,
                severity=issue.severity,
                field_name=issue.field_name,
                message_ar=issue.message_ar,
                masked_value=issue.masked_value[:255],
            )
            for row in import_rows
            for issue in issues_by_row_id[row.id]
        ]
    )


def preview_employee_import(uploaded_file, *, uploaded_by) -> EmployeeImportBatch:
    """Validate and persist an immutable preview without changing Employee data."""

    ensure_crypto_configured()
    validated = validate_xlsx_upload(uploaded_file)
    if EmployeeImportBatch.objects.filter(file_sha256=validated.file_sha256).exists():
        raise DuplicateImportFileError(code="duplicate_file")
    raw_rows = _read_rows(validated.content)
    prepared_rows = [_prepare_row(number, payload) for number, payload in raw_rows]
    effective_date = timezone.localdate()
    _add_reference_issues(prepared_rows, effective_date)
    _add_identity_and_manager_issues(prepared_rows)

    batch_id = uuid.uuid4()
    storage_key = random_storage_key(batch_id)
    encrypted_file = encrypt_sensitive_bytes(
        validated.content,
        context=f"employee-import-file:{batch_id}",
    )
    saved_storage_key = ""
    try:
        saved_storage_key = default_storage.save(
            storage_key,
            ContentFile(encrypted_file.ciphertext),
        )
        with transaction.atomic():
            batch = EmployeeImportBatch.objects.create(
                id=batch_id,
                original_filename=validated.original_filename,
                storage_key=saved_storage_key,
                file_sha256=validated.file_sha256,
                file_size_bytes=validated.file_size_bytes,
                mime_type=validated.mime_type,
                encryption_key_version=encrypted_file.key_version,
                status=EmployeeImportBatch.Status.UPLOADED,
                uploaded_by=uploaded_by,
            )
            _save_preview_rows(batch, prepared_rows)
            if not prepared_rows:
                EmployeeImportError.objects.create(
                    batch=batch,
                    error_code="no_data_rows",
                    severity=EmployeeImportError.Severity.ERROR,
                    field_name="",
                    message_ar="لا يحتوي الملف على صفوف بيانات.",
                )

            batch.total_rows = len(prepared_rows)
            batch.new_rows = sum(
                row.matched_employee is None and not row.has_errors
                for row in prepared_rows
            )
            batch.update_rows = sum(
                row.matched_employee is not None and not row.has_errors
                for row in prepared_rows
            )
            batch.missing_department_rows = sum(
                any(issue.error_code == "missing_department" for issue in row.issues)
                for row in prepared_rows
            )
            batch.missing_location_rows = sum(
                any(issue.error_code == "missing_location" for issue in row.issues)
                for row in prepared_rows
            )
            batch.unmatched_manager_rows = sum(
                any(
                    issue.error_code
                    in {
                        "invalid_manager_national_id",
                        "manager_identifier_missing",
                        "manager_not_found",
                        "manager_self_reference",
                    }
                    for issue in row.issues
                )
                for row in prepared_rows
            )
            batch.error_rows = sum(row.has_errors for row in prepared_rows)
            if not prepared_rows:
                batch.error_rows = 1
            batch.warning_rows = sum(row.has_warnings for row in prepared_rows)
            batch.status = (
                EmployeeImportBatch.Status.HAS_ERRORS
                if batch.error_rows
                else EmployeeImportBatch.Status.PREVIEW_READY
            )
            batch.save(
                update_fields=(
                    "total_rows",
                    "new_rows",
                    "update_rows",
                    "missing_department_rows",
                    "missing_location_rows",
                    "unmatched_manager_rows",
                    "error_rows",
                    "warning_rows",
                    "status",
                )
            )
            return batch
    except EmployeeImportServiceError:
        if saved_storage_key:
            default_storage.delete(saved_storage_key)
        raise
    except IntegrityError as exc:
        if saved_storage_key:
            default_storage.delete(saved_storage_key)
        if EmployeeImportBatch.objects.filter(file_sha256=validated.file_sha256).exists():
            raise DuplicateImportFileError(code="duplicate_file") from exc
        raise ImportPreviewError(code="preview_database_error") from exc
    except Exception as exc:
        if saved_storage_key:
            default_storage.delete(saved_storage_key)
        raise ImportPreviewError(code="preview_failed") from exc


def _decrypt_import_row(row: EmployeeImportRow) -> dict[str, str]:
    try:
        plaintext = decrypt_sensitive_text(
            bytes(row.raw_payload_encrypted),
            context=f"employee-import-row:{row.id}",
            key_version=row.encryption_key_version,
        )
        payload = json.loads(plaintext)
    except (TypeError, ValueError, json.JSONDecodeError) as exc:
        raise ImportApprovalError(
            "تعذر التحقق من سلامة بيانات المعاينة.",
            code="preview_integrity_error",
        ) from exc
    if not isinstance(payload, dict) or not all(
        isinstance(key, str) and isinstance(value, str)
        for key, value in payload.items()
    ):
        raise ImportApprovalError(
            "تعذر التحقق من سلامة بيانات المعاينة.",
            code="preview_integrity_error",
        )
    canonical_payload = json.dumps(
        payload,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    )
    if not hashlib.sha256(canonical_payload.encode("utf-8")).hexdigest() == row.payload_sha256:
        raise ImportApprovalError(
            "تعذر التحقق من سلامة بيانات المعاينة.",
            code="preview_integrity_error",
        )
    return payload


def _resolve_department(
    name: str,
    *,
    effective_date: date,
    approved_by,
    create_missing: bool,
) -> tuple[Department, bool]:
    matches = list(_active_departments(name, effective_date).select_for_update()[:2])
    if len(matches) > 1:
        raise ImportApprovalError(
            "تعذرت مطابقة أحد الأقسام بصورة فريدة.",
            code="ambiguous_department",
        )
    if matches:
        return matches[0], False
    if not create_missing:
        raise ImportApprovalError(
            "توجد أقسام غير موجودة. فعّل خيار إنشاء المراجع بعد مراجعتها.",
            code="missing_department",
        )
    return (
        Department.objects.create(
            code=f"IMP-DEPT-{uuid.uuid4().hex[:8].upper()}",
            name_ar=name,
            unit_type=Department.UnitType.DEPARTMENT,
            valid_from=effective_date,
            created_by=approved_by,
            updated_by=approved_by,
        ),
        True,
    )


def _resolve_location(
    name: str,
    *,
    department: Department,
    approved_by,
    create_missing: bool,
) -> tuple[Location, bool]:
    matches = list(_active_locations(name).select_for_update()[:2])
    if len(matches) > 1:
        raise ImportApprovalError(
            "تعذرت مطابقة أحد المواقع بصورة فريدة.",
            code="ambiguous_location",
        )
    if matches:
        return matches[0], False
    if not create_missing:
        raise ImportApprovalError(
            "توجد مواقع غير موجودة. فعّل خيار إنشاء المراجع بعد مراجعتها.",
            code="missing_location",
        )
    return (
        Location.objects.create(
            code=f"IMP-LOC-{uuid.uuid4().hex[:8].upper()}",
            name_ar=name,
            location_type=Location.LocationType.FIELD,
            department=department,
            created_by=approved_by,
            updated_by=approved_by,
        ),
        True,
    )


def _upsert_employee(prepared: PreparedRow, *, approved_by) -> tuple[Employee, bool]:
    if not prepared.national_id or not prepared.national_id_hash:
        raise ImportApprovalError(code="invalid_national_id")
    identity = (
        EmployeeIdentity.objects.select_for_update()
        .select_related("employee")
        .filter(national_id_hash=prepared.national_id_hash)
        .first()
    )
    masked_mobile = mask_mobile(prepared.mobile)
    if identity is not None:
        employee = identity.employee
        update_fields = ["full_name_ar", "updated_by", "updated_at"]
        employee.full_name_ar = prepared.full_name
        employee.updated_by = approved_by
        if prepared.mobile is not None:
            employee.mobile_masked = masked_mobile
            update_fields.append("mobile_masked")
        employee.save(update_fields=update_fields)
        return employee, False

    employee = Employee.objects.create(
        employee_number=None,
        full_name_ar=prepared.full_name,
        mobile_masked=masked_mobile,
        employment_status=Employee.EmploymentStatus.ACTIVE,
        created_by=approved_by,
        updated_by=approved_by,
    )
    encrypted = encrypt_sensitive_text(
        prepared.national_id,
        context=f"employee-national-id:{employee.id}",
    )
    EmployeeIdentity.objects.create(
        employee=employee,
        national_id_hash=prepared.national_id_hash,
        national_id_encrypted=encrypted.ciphertext,
        encryption_key_version=encrypted.key_version,
        national_id_last4=prepared.national_id[-4:],
        normalized_length=10,
        verified_at=timezone.now(),
        verification_source=EmployeeIdentity.VerificationSource.IMPORT,
        created_by=approved_by,
        updated_by=approved_by,
    )
    return employee, True


def _set_assignment(
    employee: Employee,
    department: Department,
    manager: Employee | None,
    *,
    effective_date: date,
    approved_by,
) -> bool:
    active = list(
        EmploymentAssignment.objects.select_for_update()
        .filter(
            employee=employee,
            is_primary=True,
            valid_from__lte=effective_date,
        )
        .filter(Q(valid_to__isnull=True) | Q(valid_to__gt=effective_date))[:2]
    )
    if len(active) > 1:
        raise ImportApprovalError(
            "توجد إسنادات وظيفية متداخلة لأحد الموظفين.",
            code="overlapping_assignments",
        )
    current = active[0] if active else None
    if current and current.department_id == department.id:
        if manager is None or current.manager_employee_id == manager.id:
            return False
        current.manager_employee = manager
        current.updated_by = approved_by
        current.reason = "تحديث المدير من استيراد بيانات الموظفين"
        current.save(
            update_fields=(
                "manager_employee",
                "updated_by",
                "reason",
                "updated_at",
            )
        )
        return True
    if current and current.valid_from == effective_date:
        current.department = department
        current.manager_employee = manager
        current.updated_by = approved_by
        current.reason = "اعتماد استيراد بيانات الموظفين"
        current.save(
            update_fields=(
                "department",
                "manager_employee",
                "updated_by",
                "reason",
                "updated_at",
            )
        )
        return True
    if current:
        current.valid_to = effective_date
        current.updated_by = approved_by
        current.save(update_fields=("valid_to", "updated_by", "updated_at"))
    EmploymentAssignment.objects.create(
        employee=employee,
        department=department,
        manager_employee=manager,
        assignment_type=EmploymentAssignment.AssignmentType.PRIMARY,
        valid_from=effective_date,
        is_primary=True,
        reason="اعتماد استيراد بيانات الموظفين",
        created_by=approved_by,
        updated_by=approved_by,
    )
    return True


def _set_primary_location(
    employee: Employee,
    location: Location,
    *,
    effective_date: date,
    approved_by,
) -> bool:
    active = list(
        EmployeePrimaryLocation.objects.select_for_update()
        .filter(employee=employee, valid_from__lte=effective_date)
        .filter(Q(valid_to__isnull=True) | Q(valid_to__gt=effective_date))[:2]
    )
    if len(active) > 1:
        raise ImportApprovalError(
            "توجد مواقع أساسية متداخلة لأحد الموظفين.",
            code="overlapping_primary_locations",
        )
    current = active[0] if active else None
    if current and current.location_id == location.id:
        return False
    if current and current.valid_from == effective_date:
        current.location = location
        current.updated_by = approved_by
        current.assignment_reason = "اعتماد استيراد بيانات الموظفين"
        current.save(
            update_fields=(
                "location",
                "updated_by",
                "assignment_reason",
                "updated_at",
            )
        )
        return True
    if current:
        current.valid_to = effective_date
        current.updated_by = approved_by
        current.save(update_fields=("valid_to", "updated_by", "updated_at"))
    EmployeePrimaryLocation.objects.create(
        employee=employee,
        location=location,
        valid_from=effective_date,
        assignment_reason="اعتماد استيراد بيانات الموظفين",
        created_by=approved_by,
        updated_by=approved_by,
    )
    return True


def _audit_approval(
    batch: EmployeeImportBatch,
    *,
    approved_by,
    created_employees: int,
    updated_employees: int,
    created_departments: int,
    created_locations: int,
    changed_assignments: int,
    changed_primary_locations: int,
) -> None:
    username = redact_potential_national_ids(
        getattr(approved_by, "username", "") or ""
    )
    AuditLog.objects.create(
        actor_user=approved_by,
        actor_username_snapshot=username[:150] or None,
        action="employee_import.approve",
        module="organization",
        object_type="EmployeeImportBatch",
        object_id=batch.id,
        object_repr_masked="دفعة استيراد موظفين",
        after_json={
            "created_employees": created_employees,
            "updated_employees": updated_employees,
            "created_departments": created_departments,
            "created_locations": created_locations,
            "changed_assignments": changed_assignments,
            "changed_primary_locations": changed_primary_locations,
        },
        reason="اعتماد استيراد بيانات الموظفين",
        outcome=AuditLog.Outcome.SUCCESS,
    )


def approve_employee_import(
    batch: EmployeeImportBatch,
    *,
    approved_by,
    create_missing_references: bool = False,
) -> EmployeeImportBatch:
    """Apply a preview once, atomically, while preserving assignment history."""

    try:
        with transaction.atomic():
            locked_batch = EmployeeImportBatch.objects.select_for_update().get(pk=batch.pk)
            if locked_batch.status == EmployeeImportBatch.Status.APPROVED:
                return locked_batch
            if locked_batch.status not in {
                EmployeeImportBatch.Status.PREVIEW_READY,
                EmployeeImportBatch.Status.HAS_ERRORS,
            }:
                raise ImportApprovalError(
                    "حالة دفعة الاستيراد لا تسمح بالاعتماد.",
                    code="invalid_batch_status",
                )
            if locked_batch.errors.filter(
                severity=EmployeeImportError.Severity.ERROR
            ).exists():
                raise ImportApprovalError(
                    "تحتوي الدفعة على أخطاء مانعة يجب معالجتها أولًا.",
                    code="blocking_preview_errors",
                )

            ensure_crypto_configured()
            import_rows = list(
                locked_batch.rows.select_for_update().order_by("row_number")
            )
            if not import_rows:
                raise ImportApprovalError(
                    "لا تحتوي الدفعة على صفوف قابلة للاعتماد.",
                    code="empty_batch",
                )
            prepared_rows = [
                _prepare_row(row.row_number, _decrypt_import_row(row))
                for row in import_rows
            ]
            if any(row.has_errors for row in prepared_rows):
                raise ImportApprovalError(
                    "تعذر التحقق من أحد صفوف المعاينة.",
                    code="row_revalidation_failed",
                )

            effective_date = timezone.localdate()
            departments: dict[str, Department] = {}
            locations: dict[tuple[str, str], Location] = {}
            created_departments = 0
            created_locations = 0
            for prepared in prepared_rows:
                department, department_created = _resolve_department(
                    prepared.department_name,
                    effective_date=effective_date,
                    approved_by=approved_by,
                    create_missing=create_missing_references,
                )
                departments[prepared.department_name] = department
                created_departments += int(department_created)
                location_key = (prepared.location_name, prepared.department_name)
                if location_key not in locations:
                    location, location_created = _resolve_location(
                        prepared.location_name,
                        department=department,
                        approved_by=approved_by,
                        create_missing=create_missing_references,
                    )
                    locations[location_key] = location
                    created_locations += int(location_created)

            employees_by_hash: dict[str, Employee] = {}
            created_employees = 0
            updated_employees = 0
            for prepared in prepared_rows:
                employee, created = _upsert_employee(prepared, approved_by=approved_by)
                employees_by_hash[prepared.national_id_hash or ""] = employee
                created_employees += int(created)
                updated_employees += int(not created)

            manager_hashes = {
                prepared.manager_national_id_hash
                for prepared in prepared_rows
                if prepared.manager_national_id_hash
            }
            existing_managers = {
                identity.national_id_hash: identity.employee
                for identity in EmployeeIdentity.objects.select_related("employee").filter(
                    national_id_hash__in=manager_hashes
                )
            }
            changed_assignments = 0
            changed_locations = 0
            for prepared in prepared_rows:
                employee = employees_by_hash[prepared.national_id_hash or ""]
                manager = None
                if prepared.manager_national_id_hash:
                    manager = employees_by_hash.get(
                        prepared.manager_national_id_hash
                    ) or existing_managers.get(prepared.manager_national_id_hash)
                    if manager is not None and manager.id == employee.id:
                        manager = None
                department = departments[prepared.department_name]
                location = locations[
                    (prepared.location_name, prepared.department_name)
                ]
                changed_assignments += int(
                    _set_assignment(
                        employee,
                        department,
                        manager,
                        effective_date=effective_date,
                        approved_by=approved_by,
                    )
                )
                changed_locations += int(
                    _set_primary_location(
                        employee,
                        location,
                        effective_date=effective_date,
                        approved_by=approved_by,
                    )
                )

            approved_at = timezone.now()
            locked_batch.status = EmployeeImportBatch.Status.APPROVED
            locked_batch.approved_by = approved_by
            locked_batch.approved_at = approved_at
            locked_batch.save(update_fields=("status", "approved_by", "approved_at"))
            _audit_approval(
                locked_batch,
                approved_by=approved_by,
                created_employees=created_employees,
                updated_employees=updated_employees,
                created_departments=created_departments,
                created_locations=created_locations,
                changed_assignments=changed_assignments,
                changed_primary_locations=changed_locations,
            )
            return locked_batch
    except EmployeeImportServiceError:
        raise
    except EmployeeImportBatch.DoesNotExist as exc:
        raise ImportApprovalError(
            "دفعة الاستيراد المطلوبة غير موجودة.",
            code="batch_not_found",
        ) from exc
    except Exception as exc:
        raise ImportApprovalError(code="approval_failed") from exc


def can_delete_employee_import(batch: EmployeeImportBatch) -> bool:
    """Allow deletion only when no approval or applied-data evidence exists."""

    if batch.status not in {
        EmployeeImportBatch.Status.UPLOADED,
        EmployeeImportBatch.Status.PREVIEW_READY,
        EmployeeImportBatch.Status.HAS_ERRORS,
        EmployeeImportBatch.Status.FAILED,
    }:
        return False
    if batch.approved_at is not None or batch.approved_by_id is not None:
        return False
    return not AuditLog.objects.filter(
        action="employee_import.approve",
        object_type="EmployeeImportBatch",
        object_id=batch.id,
        outcome=AuditLog.Outcome.SUCCESS,
    ).exists()


def delete_employee_import(batch: EmployeeImportBatch, *, deleted_by) -> None:
    """Delete one non-approved import batch and its staged data atomically."""

    try:
        with transaction.atomic():
            locked_batch = EmployeeImportBatch.objects.select_for_update().get(
                pk=batch.pk
            )
            if not can_delete_employee_import(locked_batch):
                raise ImportDeletionError(
                    (
                        "لا يمكن حذف دفعة معتمدة أو دفعة أنشأت بيانات فعلية. "
                        "تبقى هذه الدفعة محفوظة لحماية السجل التاريخي."
                    ),
                    code="approved_batch_delete_forbidden",
                )

            row_count = locked_batch.rows.count()
            error_count = locked_batch.errors.count()
            storage_key = locked_batch.storage_key
            batch_id = locked_batch.id
            status = locked_batch.status

            AuditLog.objects.create(
                actor_user=deleted_by,
                actor_username_snapshot=redact_potential_national_ids(
                    getattr(deleted_by, "username", "") or ""
                )[:150]
                or None,
                action="employee_import.delete",
                module="organization",
                object_type="EmployeeImportBatch",
                object_id=batch_id,
                object_repr_masked="دفعة استيراد موظفين تجريبية",
                before_json={
                    "status": status,
                    "preview_rows": row_count,
                    "validation_items": error_count,
                },
                after_json={"deleted": True},
                reason="حذف دفعة استيراد تجريبية غير معتمدة",
                outcome=AuditLog.Outcome.SUCCESS,
            )

            EmployeeImportError.objects.filter(batch=locked_batch).delete()
            EmployeeImportRow.objects.filter(batch=locked_batch).delete()
            locked_batch.delete()
            if storage_key:
                default_storage.delete(storage_key)
    except EmployeeImportServiceError:
        raise
    except EmployeeImportBatch.DoesNotExist as exc:
        raise ImportDeletionError(
            "دفعة الاستيراد المطلوبة غير موجودة.",
            code="batch_not_found",
        ) from exc
    except Exception as exc:
        raise ImportDeletionError(code="deletion_failed") from exc


def build_employee_import_template() -> bytes:
    """Return a small, formula-free Arabic XLSX template."""

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "الموظفون"
    worksheet.sheet_view.rightToLeft = True
    worksheet.freeze_panes = "A2"
    fill = PatternFill(fill_type="solid", fgColor="1E3A5F")
    for column, header in enumerate(EMPLOYEE_IMPORT_HEADERS, start=1):
        cell = worksheet.cell(row=1, column=column, value=header)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        worksheet.column_dimensions[cell.column_letter].width = max(18, len(header) + 5)
    worksheet.auto_filter.ref = f"A1:G1"
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()
