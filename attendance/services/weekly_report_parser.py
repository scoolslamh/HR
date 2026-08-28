from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from datetime import date, datetime, time, timedelta
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from organization.services.identity import normalize_national_id


_DIGITS = str.maketrans("٠١٢٣٤٥٦٧٨٩۰۱۲۳۴۵۶۷۸۹", "01234567890123456789")
_NULL_VALUES = {"", "-", "—", "–", "null", "none", "لا يوجد"}

HEADER_ALIASES = {
    "national_id": {"السجل المدني", "رقم السجل المدني", "الهوية", "رقم الهوية"},
    "employee_name": {"الاسم", "اسم الموظف", "الموظف"},
    "job_title": {"المسمى الوظيفي", "المسمى"},
    "attendance_date": {"التاريخ", "تاريخ الدوام", "تاريخ الحضور"},
    "source_status": {"حالة التحضير", "الحالة", "حالة الحضور"},
    "scheduled_duration": {"ساعات الدوام", "مدة الدوام"},
    "check_in": {"وقت الحضور", "توقيت الحضور"},
    "check_in_location": {"مكان الحضور", "موقع الحضور"},
    "check_out": {"توقيت الانصراف", "وقت الانصراف"},
    "check_out_location": {"مكان الانصراف", "موقع الانصراف"},
    "actual_work_duration": {"ساعات الدوام الفعلي", "ساعات العمل الفعلي"},
    "early_departure_duration": {"انصراف مبكر", "الانصراف المبكر"},
    "shortfall_duration": {"النقص في الدوام", "نقص الدوام"},
    "early_arrival_duration": {"حضور مبكر", "الحضور المبكر"},
}


@dataclass(frozen=True, slots=True)
class ParserIssue:
    row_number: int | None
    code: str
    severity: str
    field_name: str
    message_ar: str
    masked_value: str = ""


@dataclass(slots=True)
class ParsedDailyRow:
    row_number: int
    raw_payload: dict[str, Any]
    national_id: str | None
    employee_name: str
    job_title: str
    attendance_date: date | None
    source_status: str | None
    scheduled_duration: timedelta | None
    check_in: time | None
    check_in_location: str | None
    check_out: time | None
    check_out_location: str | None
    actual_work_duration: timedelta | None
    early_departure_duration: timedelta | None
    shortfall_duration: timedelta | None
    early_arrival_duration: timedelta | None
    issues: list[ParserIssue] = field(default_factory=list)


@dataclass(frozen=True, slots=True)
class ParsedWeeklyReport:
    period_title: str
    period_start: date | None
    period_end: date | None
    source_row_count: int
    employee_count: int
    ignored_row_count: int
    summary_row_count: int
    rows: tuple[ParsedDailyRow, ...]
    issues: tuple[ParserIssue, ...]


def _text(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\n", " ").split()).strip()


def _normalized_text(value: Any) -> str | None:
    rendered = _text(value).translate(_DIGITS)
    return None if rendered.lower() in _NULL_VALUES else rendered


_ARABIC_DIACRITICS_RE = re.compile(r"[\u0610-\u061A\u064B-\u065F\u0670\u06D6-\u06ED]")
_ARABIC_CANONICAL_TRANSLATION = str.maketrans({
    "أ": "ا",
    "إ": "ا",
    "آ": "ا",
    "ٱ": "ا",
    "ى": "ي",
    "ؤ": "و",
    "ئ": "ي",
})


def _header_key(value: Any) -> str:
    rendered = _text(value).translate(_DIGITS)
    rendered = rendered.translate(_ARABIC_CANONICAL_TRANSLATION).replace("ـ", "")
    rendered = _ARABIC_DIACRITICS_RE.sub("", rendered)
    return re.sub(r"[\s:：_\-]+", "", rendered).casefold()


def _alias_lookup() -> dict[str, str]:
    return {
        _header_key(alias): field
        for field, aliases in HEADER_ALIASES.items()
        for alias in aliases
    }


def _find_header(worksheet) -> tuple[int, dict[str, int]]:
    aliases = _alias_lookup()
    for row_number, values in enumerate(
        worksheet.iter_rows(min_row=1, max_row=min(worksheet.max_row, 100), values_only=True),
        start=1,
    ):
        mapping = {}
        for index, value in enumerate(values):
            canonical = aliases.get(_header_key(value))
            if canonical:
                mapping[canonical] = index
        if {"national_id", "attendance_date", "check_in", "check_out"}.issubset(mapping):
            return row_number, mapping
    raise ValueError("تعذر العثور على صف رؤوس تقرير البصمة.")


def _serialize(value: Any) -> Any:
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, timedelta):
        return value.total_seconds()
    return value


def normalize_date(value: Any, *, epoch=None) -> date | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            converted = from_excel(value, epoch=epoch)
            return converted.date() if isinstance(converted, datetime) else converted
        except (TypeError, ValueError, OverflowError):
            return None
    rendered = _normalized_text(value)
    if not rendered:
        return None
    rendered = rendered.replace(".", "/").replace("-", "/")
    candidates = [rendered]
    token_match = re.search(r"\d{1,4}/\d{1,2}/\d{1,4}", rendered)
    if token_match and token_match.group(0) != rendered:
        candidates.append(token_match.group(0))
    for candidate in candidates:
        for fmt in ("%Y/%m/%d", "%d/%m/%Y", "%Y/%d/%m", "%d/%m/%y"):
            try:
                return datetime.strptime(candidate, fmt).date()
            except ValueError:
                continue
    return None


def normalize_time(value: Any) -> time | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, datetime):
        return value.time().replace(tzinfo=None, microsecond=0)
    if isinstance(value, time):
        return value.replace(tzinfo=None, microsecond=0)
    if isinstance(value, (int, float)) and 0 <= float(value) < 1:
        seconds = int(round(float(value) * 86400)) % 86400
        return time(seconds // 3600, (seconds % 3600) // 60, seconds % 60)
    rendered = _normalized_text(value)
    if not rendered:
        return None
    rendered = rendered.replace("ص", "AM").replace("م", "PM")
    rendered = re.sub(r"\s+", " ", rendered).strip().upper()
    for fmt in ("%H:%M:%S", "%H:%M", "%I:%M:%S %p", "%I:%M %p"):
        try:
            return datetime.strptime(rendered, fmt).time()
        except ValueError:
            continue
    return None


def normalize_duration(value: Any) -> timedelta | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, timedelta):
        return value
    if isinstance(value, datetime):
        value = value.time()
    if isinstance(value, time):
        return timedelta(hours=value.hour, minutes=value.minute, seconds=value.second)
    if isinstance(value, (int, float)):
        return timedelta(days=float(value))
    rendered = _normalized_text(value)
    if not rendered:
        return None
    match = re.fullmatch(r"(\d{1,3}):([0-5]\d)(?::([0-5]\d))?", rendered)
    if match:
        return timedelta(
            hours=int(match.group(1)),
            minutes=int(match.group(2)),
            seconds=int(match.group(3) or 0),
        )
    try:
        return timedelta(hours=float(rendered))
    except ValueError:
        return None


def _period_data(worksheet, header_row: int) -> tuple[str, date | None, date | None]:
    values = []
    dates = []
    date_pattern = re.compile(r"\d{1,4}[./-]\d{1,2}[./-]\d{1,4}")
    for row in worksheet.iter_rows(min_row=1, max_row=max(1, header_row - 1), values_only=True):
        for value in row:
            rendered = _normalized_text(value)
            if not rendered:
                continue
            values.append(rendered)
            tokens = date_pattern.findall(rendered.translate(_DIGITS))
            if tokens:
                for token in tokens:
                    parsed = normalize_date(token)
                    if parsed:
                        dates.append(parsed)
            else:
                direct = normalize_date(value, epoch=worksheet.parent.epoch)
                if direct:
                    dates.append(direct)
    title = " - ".join(values)[:255]
    if not dates:
        return title, None, None
    return title, min(dates), max(dates)


def _is_null(value: Any) -> bool:
    return _normalized_text(value) is None


def parse_weekly_report(content: bytes) -> ParsedWeeklyReport:
    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=False)
    try:
        worksheet = workbook.active
        header_row, columns = _find_header(worksheet)
        period_title, period_start, period_end = _period_data(worksheet, header_row)
        ignored = header_row
        summaries = 0
        current_id = None
        current_name = ""
        current_job_title = ""
        employee_ids = set()
        parsed_rows = []
        global_issues = []

        for row_number, cells in enumerate(
            worksheet.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1,
        ):
            values = list(cells)
            if all(_is_null(value) for value in values):
                ignored += 1
                continue
            if any((_normalized_text(value) or "").startswith("المجموع") for value in values):
                summaries += 1
                continue
            if sum(1 for value in values if _header_key(value) in _alias_lookup()) >= 4:
                ignored += 1
                continue

            payload = {
                field: _serialize(values[index] if index < len(values) else None)
                for field, index in columns.items()
            }
            row_issues = []
            raw_national_id = values[columns["national_id"]] if columns["national_id"] < len(values) else None
            supplied_id = _normalized_text(raw_national_id)
            if supplied_id:
                try:
                    current_id = normalize_national_id(raw_national_id)
                    employee_ids.add(current_id)
                    current_name = _normalized_text(values[columns.get("employee_name", -1)]) if columns.get("employee_name", -1) >= 0 else ""
                    current_job_title = _normalized_text(values[columns.get("job_title", -1)]) if columns.get("job_title", -1) >= 0 else ""
                    current_name = current_name or ""
                    current_job_title = current_job_title or ""
                except ValueError:
                    current_id = None
                    row_issues.append(ParserIssue(row_number, "invalid_national_id", "error", "national_id", "السجل المدني في بداية مجموعة الموظف غير صالح."))
            else:
                if "employee_name" in columns:
                    current_name = _normalized_text(values[columns["employee_name"]]) or current_name
                if "job_title" in columns:
                    current_job_title = _normalized_text(values[columns["job_title"]]) or current_job_title

            raw_date = values[columns["attendance_date"]] if columns["attendance_date"] < len(values) else None
            attendance_date = normalize_date(raw_date, epoch=workbook.epoch)
            if _normalized_text(raw_date) and attendance_date is None:
                row_issues.append(ParserIssue(row_number, "invalid_date", "error", "attendance_date", "تاريخ الدوام غير صالح."))
            elif attendance_date is None:
                row_issues.append(ParserIssue(row_number, "missing_date", "error", "attendance_date", "صف الدوام لا يحتوي تاريخًا صالحًا."))
            if current_id is None:
                row_issues.append(ParserIssue(row_number, "missing_employee_context", "error", "national_id", "تعذر تحديد الموظف لهذا الصف."))

            def value_for(field):
                index = columns.get(field)
                return values[index] if index is not None and index < len(values) else None

            check_in = normalize_time(value_for("check_in"))
            check_out = normalize_time(value_for("check_out"))
            if _normalized_text(value_for("check_in")) and check_in is None:
                row_issues.append(ParserIssue(row_number, "invalid_check_in", "error", "check_in", "وقت الحضور غير صالح."))
            if _normalized_text(value_for("check_out")) and check_out is None:
                row_issues.append(ParserIssue(row_number, "invalid_check_out", "error", "check_out", "وقت الانصراف غير صالح."))
            source_status = _normalized_text(value_for("source_status"))
            duration_fields = {
                "scheduled_duration": "scheduled_duration",
                "actual_work_duration": "actual_work_duration",
                "early_departure_duration": "early_departure_duration",
                "shortfall_duration": "shortfall_duration",
                "early_arrival_duration": "early_arrival_duration",
            }
            durations = {}
            for source_field, target in duration_fields.items():
                raw_duration = value_for(source_field)
                durations[target] = normalize_duration(raw_duration)
                if _normalized_text(raw_duration) and durations[target] is None:
                    row_issues.append(ParserIssue(row_number, "invalid_duration", "warning", source_field, "تعذر تفسير إحدى مدد المصدر؛ حُفظت كقيمة فارغة."))
            if attendance_date and not any((check_in, check_out, source_status, *durations.values())):
                row_issues.append(ParserIssue(row_number, "incomplete_daily_row", "error", "row", "صف الدوام ناقص ولا يحتوي حالة أو أوقات أو مددًا."))

            parsed_rows.append(
                ParsedDailyRow(
                    row_number=row_number,
                    raw_payload=payload,
                    national_id=current_id,
                    employee_name=current_name,
                    job_title=current_job_title,
                    attendance_date=attendance_date,
                    source_status=source_status,
                    scheduled_duration=durations["scheduled_duration"],
                    check_in=check_in,
                    check_in_location=_normalized_text(value_for("check_in_location")),
                    check_out=check_out,
                    check_out_location=_normalized_text(value_for("check_out_location")),
                    actual_work_duration=durations["actual_work_duration"],
                    early_departure_duration=durations["early_departure_duration"],
                    shortfall_duration=durations["shortfall_duration"],
                    early_arrival_duration=durations["early_arrival_duration"],
                    issues=row_issues,
                )
            )
        if period_start is None or period_end is None:
            global_issues.append(ParserIssue(None, "period_not_found", "error", "period", "تعذر استخراج بداية ونهاية الفترة من النص أعلى الشيت."))
        return ParsedWeeklyReport(
            period_title=period_title,
            period_start=period_start,
            period_end=period_end,
            source_row_count=worksheet.max_row,
            employee_count=len(employee_ids),
            ignored_row_count=ignored,
            summary_row_count=summaries,
            rows=tuple(parsed_rows),
            issues=tuple(global_issues),
        )
    finally:
        workbook.close()
