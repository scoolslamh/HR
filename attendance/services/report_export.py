from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO

from django.db.models import Count, Q, Sum
from openpyxl import Workbook

from attendance.models import DailyAttendanceResult


REPORT_TYPES = {
    "summary": "ملخص إحصائي",
    "comprehensive": "تقرير شامل",
    "top_absence": "الأكثر غيابًا",
    "top_permissions": "الأكثر استئذانًا",
    "top_late": "الأكثر تأخرًا",
    "top_early_leave": "الأكثر انصرافًا مبكرًا",
    "outside_location": "اختلاف مواقع التوقيع",
    "work_missions": "مهمات العمل",
    "details": "سجلات تفصيلية",
}


@dataclass(frozen=True, slots=True)
class ReportData:
    title: str
    summary: tuple[tuple[str, int | str], ...]
    columns: tuple[str, ...]
    rows: tuple[tuple, ...]
    sections: tuple["ReportSection", ...] = ()


@dataclass(frozen=True, slots=True)
class ReportSection:
    title: str
    columns: tuple[str, ...]
    rows: tuple[tuple, ...]


def _summary(queryset) -> tuple[tuple[str, int], ...]:
    aggregate = queryset.aggregate(
        records=Count("id"),
        employees=Count("employee_id", distinct=True),
        late_minutes=Sum("late_minutes"),
        early_leave_minutes=Sum("early_leave_minutes"),
    )
    return (
        ("إجمالي السجلات", aggregate["records"] or 0),
        ("عدد الموظفين", aggregate["employees"] or 0),
        ("أيام الغياب", queryset.filter(attendance_status=DailyAttendanceResult.AttendanceStatus.ABSENT).count()),
        ("الاستئذانات", queryset.filter(source_status__icontains="استئذان").count()),
        ("مهمات العمل", queryset.filter(Q(source_status__icontains="مهمة عمل") | Q(source_status__icontains="مهمه عمل")).count()),
        ("اختلاف مواقع التوقيع", queryset.filter(location_status=DailyAttendanceResult.LocationStatus.BOTH_OUTSIDE).count()),
        ("إجمالي دقائق التأخر", aggregate["late_minutes"] or 0),
        ("إجمالي دقائق الانصراف المبكر", aggregate["early_leave_minutes"] or 0),
    )


def _ranking(queryset, report_type: str, limit: int):
    metric_label = "عدد الحالات"
    if report_type == "top_absence":
        queryset = queryset.filter(attendance_status=DailyAttendanceResult.AttendanceStatus.ABSENT)
        metric = Count("id")
    elif report_type == "top_permissions":
        queryset = queryset.filter(source_status__icontains="استئذان")
        metric = Count("id")
    elif report_type == "top_late":
        queryset = queryset.filter(late_minutes__gt=0)
        metric = Sum("late_minutes")
        metric_label = "إجمالي دقائق التأخر"
    elif report_type == "top_early_leave":
        queryset = queryset.filter(early_leave_minutes__gt=0)
        metric = Sum("early_leave_minutes")
        metric_label = "إجمالي دقائق الانصراف المبكر"
    elif report_type == "outside_location":
        queryset = queryset.filter(location_status=DailyAttendanceResult.LocationStatus.BOTH_OUTSIDE)
        metric = Count("id")
    elif report_type == "work_missions":
        queryset = queryset.filter(Q(source_status__icontains="مهمة عمل") | Q(source_status__icontains="مهمه عمل"))
        metric = Count("id")
    else:
        queryset = queryset.filter(attendance_status=DailyAttendanceResult.AttendanceStatus.ABSENT)
        metric = Count("id")
    values = (
        queryset.values("employee_id", "employee__full_name_ar")
        .annotate(metric=metric)
        .order_by("-metric", "employee__full_name_ar")[:limit]
    )
    rows = tuple(
        (index, item["employee__full_name_ar"], item["metric"] or 0)
        for index, item in enumerate(values, start=1)
    )
    return ("الترتيب", "الموظف", metric_label), rows


def _discipline_filter() -> Q:
    return (
        Q(attendance_status=DailyAttendanceResult.AttendanceStatus.PRESENT)
        & Q(late_minutes=0)
        & Q(early_leave_minutes=0)
        & ~Q(location_status=DailyAttendanceResult.LocationStatus.BOTH_OUTSIDE)
    )


def _comprehensive_summary(queryset) -> tuple[tuple[str, int | str], ...]:
    aggregate = queryset.aggregate(
        records=Count("id"),
        employees=Count("employee_id", distinct=True),
        disciplined=Count("id", filter=_discipline_filter()),
    )
    total = aggregate["records"] or 0
    discipline_percentage = round(((aggregate["disciplined"] or 0) / total) * 100) if total else 0
    return (
        ("عدد الموظفين", aggregate["employees"] or 0),
        (
            "أيام الغياب",
            queryset.filter(
                attendance_status=DailyAttendanceResult.AttendanceStatus.ABSENT
            ).count(),
        ),
        ("الاستئذانات", queryset.filter(source_status__icontains="استئذان").count()),
        (
            "مهمات العمل",
            queryset.filter(
                Q(source_status__icontains="مهمة عمل")
                | Q(source_status__icontains="مهمه عمل")
            ).count(),
        ),
        ("نسبة الانضباط العامة", f"{discipline_percentage}%"),
        (
            "اختلاف مواقع التوقيع",
            queryset.filter(
                location_status=DailyAttendanceResult.LocationStatus.BOTH_OUTSIDE
            ).count(),
        ),
    )


def _discipline_ranking(queryset, limit: int = 10):
    values = queryset.values("employee_id", "employee__full_name_ar").annotate(
        total=Count("id"),
        disciplined=Count("id", filter=_discipline_filter()),
    )
    ranked = sorted(
        values,
        key=lambda item: (
            -((item["disciplined"] or 0) / item["total"]),
            -(item["disciplined"] or 0),
            item["employee__full_name_ar"],
        ),
    )[:limit]
    rows = tuple(
        (
            index,
            item["employee__full_name_ar"],
            item["disciplined"] or 0,
            item["total"],
            f"{round(((item['disciplined'] or 0) / item['total']) * 100)}%",
        )
        for index, item in enumerate(ranked, start=1)
    )
    return ("الترتيب", "الموظف", "الأيام المنضبطة", "إجمالي الأيام", "نسبة الانضباط"), rows


def _comprehensive_sections(queryset) -> tuple[ReportSection, ...]:
    sections = []
    for title, report_type in (
        ("أكثر 10 موظفين استئذانًا", "top_permissions"),
        ("أكثر 10 موظفين في مهمات العمل", "work_missions"),
        ("أكثر 10 موظفين غيابًا", "top_absence"),
    ):
        columns, rows = _ranking(queryset, report_type, 10)
        sections.append(ReportSection(title, columns, rows))
    columns, rows = _discipline_ranking(queryset)
    sections.append(ReportSection("أكثر 10 موظفين انضباطًا", columns, rows))
    return tuple(sections)


def build_report(queryset, *, report_type: str, limit: int = 10) -> ReportData:
    report_type = report_type if report_type in REPORT_TYPES else "summary"
    if report_type == "comprehensive":
        return ReportData(
            REPORT_TYPES[report_type],
            _comprehensive_summary(queryset),
            (),
            (),
            _comprehensive_sections(queryset),
        )
    summary = _summary(queryset)
    if report_type == "summary":
        return ReportData(REPORT_TYPES[report_type], summary, (), ())
    if report_type == "details":
        values = queryset.order_by("-attendance_date", "employee__full_name_ar").values_list(
            "employee__full_name_ar", "department__name_ar", "attendance_date",
            "source_status", "worked_minutes", "late_minutes",
            "early_leave_minutes", "shortfall_minutes",
        )[:limit]
        return ReportData(
            REPORT_TYPES[report_type], summary,
            ("الموظف", "القسم", "التاريخ", "حالة المصدر", "دقائق العمل", "دقائق التأخر", "دقائق الانصراف المبكر", "دقائق النقص"),
            tuple(values),
        )
    columns, rows = _ranking(queryset, report_type, limit)
    return ReportData(REPORT_TYPES[report_type], summary, columns, rows)


def report_xlsx(report: ReportData) -> bytes:
    output = BytesIO()
    workbook = Workbook(write_only=True)
    summary_sheet = workbook.create_sheet("الملخص")
    summary_sheet.sheet_view.rightToLeft = True
    summary_sheet.append((report.title, ""))
    summary_sheet.append(("المؤشر", "القيمة"))
    for item in report.summary:
        summary_sheet.append(item)
    if report.columns:
        detail_sheet = workbook.create_sheet("البيانات")
        detail_sheet.sheet_view.rightToLeft = True
        detail_sheet.append(report.columns)
        for row in report.rows:
            detail_sheet.append(row)
    for section in report.sections:
        section_sheet = workbook.create_sheet(section.title[:31])
        section_sheet.sheet_view.rightToLeft = True
        section_sheet.append(section.columns)
        for row in section.rows:
            section_sheet.append(row)
    workbook.save(output)
    workbook.close()
    return output.getvalue()
