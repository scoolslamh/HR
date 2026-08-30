import time as time_module
import tracemalloc
from datetime import date, datetime, time, timedelta
from io import BytesIO

from openpyxl import load_workbook

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.db import connection
from django.test.utils import CaptureQueriesContext
from django.urls import reverse
from django.utils import timezone

from accounts.models import Permission, Role, RolePermission, UserRole

from organization.models import (
    Department,
    Employee,
    EmployeeIdentity,
    EmployeePrimaryLocation,
    EmploymentAssignment,
    Location,
    UserDepartmentScope,
)
from organization.services.identity import (
    encrypt_sensitive_text,
    national_id_digest,
)

from attendance.models import (
    CalculationRun,
    DailyAttendanceResult,
    ImportBatch,
    ImportRow,
    RawAttendanceRecord,
)
from attendance.services.calculation import calculate_records
from attendance.services.weekly_import import (
    archive_attendance_import,
    restore_attendance_import,
    update_attendance_import_metadata,
)
from violations.models import ClarificationRequest


class AttendanceCalculationTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_superuser(
            username="calc-admin", password="Strong-Test-Pass-2026"
        )
        self.department = Department.objects.create(
            code="D-1",
            name_ar="إدارة الاختبار",
            unit_type=Department.UnitType.DEPARTMENT,
            valid_from=date(2026, 1, 1),
        )
        self.location = Location.objects.create(
            code="L-1",
            name_ar="المبنى الرئيسي",
            location_type=Location.LocationType.HEADQUARTERS,
        )
        self.employee = Employee.objects.create(full_name_ar="موظف الاختبار")
        self.national_id = "1023456789"
        encrypted_national_id = encrypt_sensitive_text(
            self.national_id,
            context=f"employee-national-id:{self.employee.id}",
        )
        EmployeeIdentity.objects.create(
            employee=self.employee,
            national_id_hash=national_id_digest(self.national_id),
            national_id_encrypted=encrypted_national_id.ciphertext,
            encryption_key_version=encrypted_national_id.key_version,
            national_id_last4=self.national_id[-4:],
        )
        EmploymentAssignment.objects.create(
            employee=self.employee,
            department=self.department,
            valid_from=date(2026, 1, 1),
            is_primary=True,
        )
        EmployeePrimaryLocation.objects.create(
            employee=self.employee,
            location=self.location,
            valid_from=date(2026, 1, 1),
        )
        self.batch = ImportBatch.objects.create(
            original_filename="weekly.xlsx",
            storage_key="attendance/imports/test.bin",
            file_sha256="a" * 64,
            file_size_bytes=100,
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            period_start=date(2026, 6, 28),
            period_end=date(2026, 7, 2),
            status=ImportBatch.Status.APPROVED,
            approved_by=self.user,
            approved_at=timezone.now(),
        )

    def _record(
        self,
        *,
        in_location="المبنى الرئيسي",
        out_location="المبنى الرئيسي",
        attendance_date=date(2026, 6, 28),
        row_number=1,
        source_status="مكتملة",
    ):
        row = ImportRow.objects.create(
            batch=self.batch,
            row_number=row_number,
            raw_payload_encrypted=b"encrypted",
            encryption_key_version="v1",
            raw_payload_sha256=f"{row_number:064x}",
            normalized_payload_json={},
            display_data_json={},
            matched_employee=self.employee,
            attendance_date=attendance_date,
            match_status=ImportRow.MatchStatus.MATCHED,
            validation_status=ImportRow.ValidationStatus.VALID,
            location_match_status=ImportRow.LocationMatchStatus.MATCHED,
        )
        return RawAttendanceRecord.objects.create(
            import_row=row,
            employee=self.employee,
            national_id_hash="c" * 64,
            attendance_date=attendance_date,
            source_check_in_at=timezone.make_aware(datetime.combine(attendance_date, time(7, 10))),
            source_check_out_at=timezone.make_aware(datetime.combine(attendance_date, time(13, 40))),
            source_check_in_location=in_location,
            source_check_out_location=out_location,
            primary_location=self.location,
            source_status=source_status,
            source_scheduled_duration=timedelta(hours=7),
            source_actual_work_duration=timedelta(hours=6, minutes=30),
            source_early_departure_duration=timedelta(minutes=10),
            source_shortfall_duration=timedelta(minutes=30),
            source_early_arrival_duration=timedelta(0),
            record_fingerprint=f"{row_number + 100:064x}",
            location_match_status=ImportRow.LocationMatchStatus.MATCHED,
            matched_at=timezone.now(),
        )

    def test_calculates_daily_result_and_department_snapshot(self):
        record = self._record()
        summary = calculate_records(records=[record], requested_by=self.user, import_batch=self.batch)

        self.assertEqual(summary.created, 1)
        result = DailyAttendanceResult.objects.get()
        self.assertEqual(result.department, self.department)
        self.assertEqual(result.worked_minutes, 390)
        self.assertEqual(result.shortfall_minutes, 30)
        self.assertEqual(result.early_leave_minutes, 10)
        self.assertEqual(result.late_minutes, 20)
        self.assertEqual(result.location_status, DailyAttendanceResult.LocationStatus.MATCHED)
        self.assertEqual(result.attendance_status, DailyAttendanceResult.AttendanceStatus.PRESENT)

    def test_calculates_15000_records_in_bounded_batches(self):
        row_count = 15_000
        start = date(2030, 1, 1)
        rows = []
        records = []
        for index in range(row_count):
            attendance_day = start + timedelta(days=index)
            row = ImportRow(
                batch=self.batch,
                row_number=index + 1,
                raw_payload_encrypted=b"encrypted",
                encryption_key_version="v1",
                raw_payload_sha256=f"{index + 1:064x}",
                normalized_payload_json={},
                display_data_json={},
                matched_employee=self.employee,
                attendance_date=attendance_day,
                match_status=ImportRow.MatchStatus.MATCHED,
                validation_status=ImportRow.ValidationStatus.VALID,
                location_match_status=ImportRow.LocationMatchStatus.UNKNOWN,
            )
            rows.append(row)
        ImportRow.objects.bulk_create(rows, batch_size=500)
        for index, row in enumerate(rows):
            records.append(
                RawAttendanceRecord(
                    import_row=row,
                    employee=self.employee,
                    national_id_hash="c" * 64,
                    attendance_date=row.attendance_date,
                    source_status="غياب",
                    record_fingerprint=f"{index + 100_000:064x}",
                    location_match_status=ImportRow.LocationMatchStatus.UNKNOWN,
                    matched_at=timezone.now(),
                )
            )
        RawAttendanceRecord.objects.bulk_create(records, batch_size=500)
        rows.clear()
        records.clear()

        tracemalloc.start()
        started = time_module.perf_counter()
        with CaptureQueriesContext(connection) as queries:
            summary = calculate_records(
                records=RawAttendanceRecord.objects.filter(import_row__batch=self.batch),
                requested_by=self.user,
                import_batch=self.batch,
            )
        elapsed = time_module.perf_counter() - started
        _, peak_bytes = tracemalloc.get_traced_memory()
        tracemalloc.stop()

        self.assertEqual(summary.created, row_count)
        self.assertEqual(DailyAttendanceResult.objects.count(), row_count)
        self.assertEqual(ClarificationRequest.objects.count(), row_count)
        print(
            f"15k calculation: {len(queries)} queries, {elapsed:.2f}s, "
            f"Python peak {peak_bytes / 1024 / 1024:.1f} MiB"
        )
        # SQLite splits bulk statements at its low bind-parameter limit;
        # PostgreSQL uses substantially fewer statements for the same batches.
        self.assertLess(len(queries), 1_500)
        self.assertLess(peak_bytes, 160 * 1024 * 1024)

    def test_system_admin_can_manage_and_archive_batch_from_import_page(self):
        record = self._record()
        calculate_records(records=[record], requested_by=self.user, import_batch=self.batch)
        update_attendance_import_metadata(
            self.batch,
            display_name="حضور الأسبوع التجريبي",
            source_period_title="الأسبوع الأخير من يونيو",
            reason="تنظيم اسم الملف للفترة",
            updated_by=self.user,
        )
        archive_attendance_import(
            self.batch,
            archived_by=self.user,
            reason="استبعاد الملف مؤقتًا للمراجعة",
        )
        self.batch.refresh_from_db()
        self.assertEqual(self.batch.display_name, "حضور الأسبوع التجريبي")
        self.assertIsNotNone(self.batch.archived_at)

        self.client.force_login(self.user)
        archived_records = self.client.get(reverse("attendance:record_list"))
        self.assertNotContains(archived_records, "موظف الاختبار")
        import_page = self.client.get(reverse("attendance:import_list"))
        self.assertContains(import_page, "حضور الأسبوع التجريبي")
        self.assertContains(import_page, "مؤرشف")

        restore_attendance_import(
            self.batch,
            restored_by=self.user,
            reason="انتهاء مراجعة الملف واستعادته",
        )
        restored_records = self.client.get(reverse("attendance:record_list"))
        self.assertContains(restored_records, "موظف الاختبار")

        importer = get_user_model().objects.create_user(
            username="attendance-importer", password="Strong-Test-Pass-2026"
        )
        importer_role = Role.objects.create(code="importer-test", name_ar="مستورد اختبار")
        RolePermission.objects.create(
            role=importer_role,
            permission=Permission.objects.get(code="attendance.import"),
        )
        UserRole.objects.create(
            user=importer,
            role=importer_role,
            valid_from=timezone.now(),
        )
        self.client.force_login(importer)
        importer_page = self.client.get(reverse("attendance:import_list"))
        self.assertEqual(importer_page.status_code, 200)
        self.assertNotContains(
            importer_page,
            reverse("attendance:import_update", args=(self.batch.id,)),
        )
        forbidden_update = self.client.post(
            reverse("attendance:import_update", args=(self.batch.id,)),
            {
                "display_name": "غير مسموح",
                "source_period_title": "",
                "reason": "محاولة غير مسموحة",
            },
        )
        self.assertEqual(forbidden_update.status_code, 403)

    def test_system_admin_can_delete_unapproved_batch_with_confirmation(self):
        draft = ImportBatch.objects.create(
            original_filename="draft.xlsx",
            display_name="مسودة حضور أغسطس",
            storage_key="attendance/imports/missing-test-file.bin",
            file_sha256="d" * 64,
            file_size_bytes=100,
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            status=ImportBatch.Status.PREVIEW_READY,
            uploaded_by=self.user,
        )
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("attendance:import_delete", args=(draft.id,)),
            {
                "confirmation": "مسودة حضور أغسطس",
                "reason": "حذف مسودة غير معتمدة بعد المراجعة",
            },
        )

        self.assertRedirects(response, reverse("attendance:import_list"))
        self.assertFalse(ImportBatch.objects.filter(pk=draft.id).exists())

    def test_department_head_dashboard_uses_real_scoped_attendance(self):
        record = self._record()
        calculate_records(records=[record], requested_by=self.user, import_batch=self.batch)
        department_head = get_user_model().objects.create_user(
            username="department-head",
            password="Strong-Test-Pass-2026",
            first_name="رئيس",
            last_name="القسم",
        )
        UserDepartmentScope.objects.create(
            user=department_head,
            department=self.department,
            access_level=UserDepartmentScope.AccessLevel.VIEW,
            valid_from=timezone.now() - timedelta(days=1),
        )
        self.client.force_login(department_head)

        response = self.client.get(reverse("core:dashboard"))

        self.assertEqual(response.status_code, 200)
        stats = {item["title"]: item["value"] for item in response.context["dashboard_stats"]}
        self.assertEqual(stats["عدد الموظفين"], 1)
        self.assertEqual(stats["أيام الغياب"], 0)
        self.assertEqual(stats["الاستئذانات"], 0)
        self.assertContains(response, "نسبة انضباط فرق العمل")
        self.assertNotContains(response, "رسم الحضور الأسبوعي")
        self.assertNotContains(response, "بيانات توضيحية")
        self.assertNotContains(response, "weekly.xlsx")

    def test_department_head_work_mission_list_matches_dashboard_scope_when_snapshot_is_missing(self):
        mission = self._record(source_status="مهمة عمل رسمية")
        calculate_records(
            records=[mission], requested_by=self.user, import_batch=self.batch
        )
        DailyAttendanceResult.objects.update(department=None)

        head_user = get_user_model().objects.create_user(
            username="mission-scope-head",
            password="Strong-Test-Pass-2026",
        )
        head_employee = Employee.objects.create(
            full_name_ar="رئيس قسم المهمات", user=head_user
        )
        self.department.department_head = head_employee
        self.department.save(update_fields=("department_head", "updated_at"))
        UserRole.objects.create(
            user=head_user,
            role=Role.objects.get(code="department_head"),
            valid_from=timezone.now(),
        )
        self.client.force_login(head_user)

        dashboard = self.client.get(reverse("core:dashboard"))
        mission_stat = next(
            item for item in dashboard.context["dashboard_stats"]
            if item["title"] == "مهمات العمل"
        )
        mission_list = self.client.get(reverse("violations:work_mission_list"))

        self.assertEqual(mission_stat["value"], 1)
        self.assertEqual(mission_list.status_code, 200)
        self.assertEqual(mission_list.context["mission_total"], 1)
        self.assertContains(mission_list, self.employee.full_name_ar)

    def test_dashboard_uses_current_department_when_snapshot_is_missing(self):
        assignment = self.employee.employment_assignments.get(is_primary=True)
        assignment.valid_from = date(2026, 7, 12)
        assignment.save(update_fields=("valid_from", "updated_at"))
        record = self._record()
        calculate_records(
            records=[record], requested_by=self.user, import_batch=self.batch
        )
        result = DailyAttendanceResult.objects.get()
        self.assertIsNone(result.department_id)
        result.late_minutes = 0
        result.early_leave_minutes = 0
        result.save(update_fields=("late_minutes", "early_leave_minutes"))
        self.client.force_login(self.user)

        response = self.client.get(reverse("core:dashboard"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.context["unit_statuses"],
            (
                {
                    "name": self.department.name_ar,
                    "value": 100,
                    "status": "مستقر",
                    "color": "green",
                },
            ),
        )
        self.assertContains(response, self.department.name_ar)
        self.assertContains(response, "100%")

    def test_detects_different_check_in_and_check_out_locations_without_clarification(self):
        record = self._record(in_location="فرع آخر")
        calculate_records(records=[record], requested_by=self.user, import_batch=self.batch)

        result = DailyAttendanceResult.objects.get()
        self.assertFalse(result.check_in_location_matches)
        self.assertFalse(result.check_out_location_matches)
        self.assertEqual(
            result.location_status,
            DailyAttendanceResult.LocationStatus.BOTH_OUTSIDE,
        )
        self.assertFalse(
            ClarificationRequest.objects.filter(
                employee=self.employee,
                kind=ClarificationRequest.Kind.OUTSIDE_LOCATION,
            ).exists()
        )

    def test_missing_location_is_unknown_without_clarification(self):
        record = self._record(out_location="")
        calculate_records(records=[record], requested_by=self.user, import_batch=self.batch)

        result = DailyAttendanceResult.objects.get()
        self.assertIsNone(result.check_in_location_matches)
        self.assertIsNone(result.check_out_location_matches)
        self.assertEqual(result.location_status, DailyAttendanceResult.LocationStatus.UNKNOWN)
        self.assertFalse(
            ClarificationRequest.objects.filter(
                employee=self.employee,
                kind=ClarificationRequest.Kind.OUTSIDE_LOCATION,
            ).exists()
        )

    def test_record_and_outside_location_pages_render(self):
        record = self._record(in_location="فرع آخر")
        calculate_records(records=[record], requested_by=self.user, import_batch=self.batch)
        self.client.force_login(self.user)

        record_response = self.client.get(reverse("attendance:record_list"))
        report_response = self.client.get(reverse("attendance:outside_location_report"))

        self.assertEqual(record_response.status_code, 200)
        self.assertContains(record_response, "موظف الاختبار")
        self.assertEqual(report_response.status_code, 200)
        self.assertContains(report_response, "أيام اختلاف الموقعين")
        self.assertContains(report_response, "1 يوم")

        details_response = self.client.get(
            reverse("attendance:outside_location_report"),
            {"details": "outside", "details_employee": self.employee.id},
        )
        self.assertEqual(details_response.status_code, 200)
        self.assertContains(details_response, "فرع آخر")

    def test_outside_location_report_groups_employee_and_counts_days(self):
        first_record = self._record(in_location="فرع أول")
        second_record = self._record(
            in_location="فرع ثانٍ",
            attendance_date=date(2026, 6, 29),
            row_number=2,
        )
        calculate_records(
            records=[first_record, second_record],
            requested_by=self.user,
            import_batch=self.batch,
        )
        self.client.force_login(self.user)

        response = self.client.get(reverse("attendance:outside_location_report"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.context["page_obj"].object_list), 1)
        self.assertEqual(response.context["page_obj"].object_list[0]["violation_days"], 2)
        self.assertContains(response, "موظف الاختبار")
        self.assertContains(response, "2 يوم")

    def test_report_counts_imported_absence_checkout_and_permission_statuses(self):
        records = [
            self._record(
                attendance_date=date(2026, 6, 28),
                row_number=1,
                source_status="غياب",
            ),
            self._record(
                attendance_date=date(2026, 6, 29),
                row_number=2,
                source_status="انصراف تلقائي",
            ),
            self._record(
                attendance_date=date(2026, 6, 30),
                row_number=3,
                source_status="استئذان طبي",
            ),
            self._record(
                attendance_date=date(2026, 7, 1),
                row_number=4,
                source_status="إجازة سنوية",
            ),
        ]
        calculate_records(records=records, requested_by=self.user, import_batch=self.batch)
        self.client.force_login(self.user)

        response = self.client.get(reverse("attendance:outside_location_report"))
        summary = response.context["page_obj"].object_list[0]

        self.assertEqual(summary["absence_days"], 1)
        self.assertEqual(summary["automatic_checkout_days"], 1)
        self.assertEqual(summary["permission_days"], 1)
        dashboard = self.client.get(reverse("core:dashboard"))
        dashboard_stats = {
            item["title"]: item["value"]
            for item in dashboard.context["dashboard_stats"]
        }
        self.assertEqual(dashboard_stats["أيام الغياب"], 1)
        self.assertEqual(dashboard_stats["الإجازات السنوية"], 1)
        self.assertEqual(dashboard_stats["الاستئذانات"], 1)
        self.assertEqual(
            dashboard.context["top_absent_employees"][0],
            (self.employee.full_name_ar, 1),
        )
        self.assertEqual(
            dashboard.context["top_permission_employees"][0],
            (self.employee.full_name_ar, 1),
        )
        self.assertSetEqual(
            set(
                ClarificationRequest.objects.filter(employee=self.employee).values_list(
                    "kind", flat=True
                )
            ),
            {
                ClarificationRequest.Kind.ABSENCE,
                ClarificationRequest.Kind.AUTOMATIC_CHECKOUT,
            },
        )

        for details_kind, expected_status in (
            ("absent", "غياب"),
            ("automatic_checkout", "انصراف تلقائي"),
            ("permissions", "استئذان طبي"),
        ):
            details_response = self.client.get(
                reverse("attendance:outside_location_report"),
                {"details": details_kind, "details_employee": self.employee.id},
            )
            self.assertEqual(details_response.status_code, 200)
            self.assertContains(details_response, expected_status)

    def test_calculation_creates_automatic_clarifications_without_duplicates(self):
        absence = self._record(source_status="غياب")
        calculate_records(records=[absence], requested_by=self.user, import_batch=self.batch)
        calculate_records(records=[absence], requested_by=self.user, import_batch=self.batch)

        clarification = ClarificationRequest.objects.get(
            employee=self.employee,
            attendance_date=absence.attendance_date,
            kind=ClarificationRequest.Kind.ABSENCE,
        )
        self.assertEqual(ClarificationRequest.objects.count(), 1)
        self.assertEqual(clarification.attendance_result.version, 2)

    def test_report_cards_and_category_links_use_source_statuses(self):
        self.batch.period_end = date(2026, 7, 5)
        self.batch.save(update_fields=("period_end",))
        records = [
            self._record(
                attendance_date=date(2026, 6, 28),
                row_number=1,
                source_status="إجازة طارئة",
            ),
            self._record(
                attendance_date=date(2026, 6, 29),
                row_number=2,
                source_status="إجازة طبية",
            ),
            self._record(
                attendance_date=date(2026, 6, 30),
                row_number=3,
                source_status="مهمة عمل رسمية",
            ),
            self._record(
                attendance_date=date(2026, 7, 1),
                row_number=4,
                source_status="انتداب",
            ),
            self._record(
                attendance_date=date(2026, 7, 2),
                row_number=5,
                source_status="انصراف تلقائي",
            ),
            self._record(
                attendance_date=date(2026, 7, 3),
                row_number=6,
                source_status="تدريب خارجي",
            ),
            self._record(
                attendance_date=date(2026, 7, 4),
                row_number=7,
                source_status="استئذان طارئ",
            ),
            self._record(
                attendance_date=date(2026, 7, 5),
                row_number=8,
                source_status="استئذان طبي",
            ),
        ]
        calculate_records(
            records=records,
            requested_by=self.user,
            import_batch=self.batch,
        )
        self.client.force_login(self.user)

        response = self.client.get(reverse("attendance:report_overview"))

        self.assertEqual(response.context["summary"]["emergency_leave"], 1)
        self.assertEqual(response.context["summary"]["medical_leave"], 1)
        self.assertEqual(response.context["summary"]["work_missions"], 1)
        self.assertEqual(response.context["summary"]["delegation"], 1)
        self.assertEqual(response.context["summary"]["training"], 1)
        self.assertEqual(response.context["summary"]["emergency_permission"], 1)
        self.assertEqual(response.context["summary"]["medical_permission"], 1)
        self.assertEqual(response.context["summary"]["automatic_checkout"], 1)
        for removed_label in (
            "إجمالي ساعات العمل",
            "إجمالي التأخر",
            "الانصراف المبكر",
            "نقص الدوام",
        ):
            self.assertNotContains(response, removed_label)

        category_response = self.client.get(
            reverse("attendance:category_report", args=("work-missions",))
        )
        self.assertEqual(category_response.status_code, 200)
        self.assertContains(category_response, self.employee.full_name_ar)
        self.assertContains(category_response, "مهمات العمل")

    def test_report_builder_previews_and_exports_scoped_data(self):
        absence = self._record(source_status="غياب")
        calculate_records(
            records=[absence], requested_by=self.user, import_batch=self.batch
        )
        self.client.force_login(self.user)
        params = {
            "report_type": "top_absence",
            "limit": "10",
            "output_format": "preview",
        }

        preview = self.client.get(reverse("attendance:report_builder"), params)
        self.assertEqual(preview.status_code, 200)
        self.assertContains(preview, self.employee.full_name_ar)
        self.assertContains(preview, "أيام الغياب")

        params["output_format"] = "xlsx"
        exported = self.client.get(reverse("attendance:report_builder"), params)
        self.assertEqual(exported.status_code, 200)
        self.assertEqual(
            exported["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(BytesIO(exported.content), read_only=True)
        self.assertEqual(workbook.sheetnames, ["الملخص", "البيانات"])
        self.assertEqual(workbook["البيانات"]["B2"].value, self.employee.full_name_ar)
        workbook.close()

        params["output_format"] = "pdf"
        printable = self.client.get(reverse("attendance:report_builder"), params)
        self.assertEqual(printable.status_code, 200)
        self.assertContains(printable, "طباعة / حفظ PDF")
        self.assertContains(printable, self.employee.full_name_ar)

    def test_employee_report_summarizes_selected_employee_and_date_range(self):
        records = [
            self._record(
                attendance_date=date(2026, 6, 28),
                row_number=1,
                source_status="غياب",
            ),
            self._record(
                attendance_date=date(2026, 6, 29),
                row_number=2,
                source_status="استئذان طبي",
            ),
            self._record(
                attendance_date=date(2026, 6, 30),
                row_number=3,
                source_status="إجازة سنوية",
            ),
        ]
        calculate_records(
            records=records, requested_by=self.user, import_batch=self.batch
        )
        self.client.force_login(self.user)

        response = self.client.get(
            reverse("attendance:employee_report"),
            {
                "employee": self.employee.id,
                "date_from": "2026-06-28",
                "date_to": "2026-06-29",
            },
        )

        self.assertEqual(response.status_code, 200)
        summary = dict(response.context["summary"])
        self.assertEqual(summary["إجمالي الأيام"], 2)
        self.assertEqual(summary["أيام الغياب"], 1)
        self.assertEqual(summary["الاستئذانات"], 1)
        self.assertEqual(summary["الإجازات"], 0)
        self.assertEqual(len(response.context["rows"]), 2)
        self.assertContains(response, self.employee.full_name_ar)
        self.assertContains(response, "استئذان طبي")
        self.assertNotContains(response, "إجازة سنوية")

    def test_comprehensive_report_contains_executive_summary_and_four_rankings(self):
        records = [
            self._record(
                attendance_date=date(2026, 6, 28),
                row_number=1,
                source_status="غياب",
            ),
            self._record(
                attendance_date=date(2026, 6, 29),
                row_number=2,
                source_status="استئذان طبي",
            ),
            self._record(
                attendance_date=date(2026, 6, 30),
                row_number=3,
                source_status="مهمة عمل رسمية",
            ),
            self._record(
                attendance_date=date(2026, 7, 1),
                row_number=4,
                source_status="مكتملة",
            ),
        ]
        calculate_records(
            records=records, requested_by=self.user, import_batch=self.batch
        )
        DailyAttendanceResult.objects.filter(
            attendance_date=date(2026, 7, 1)
        ).update(late_minutes=0, early_leave_minutes=0)
        self.client.force_login(self.user)
        params = {
            "report_type": "comprehensive",
            "limit": "10",
            "output_format": "preview",
        }

        response = self.client.get(reverse("attendance:report_builder"), params)

        self.assertEqual(response.status_code, 200)
        report = response.context["report"]
        summary = dict(report.summary)
        self.assertEqual(summary["عدد الموظفين"], 1)
        self.assertEqual(summary["أيام الغياب"], 1)
        self.assertEqual(summary["الاستئذانات"], 1)
        self.assertEqual(summary["مهمات العمل"], 1)
        self.assertEqual(summary["نسبة الانضباط العامة"], "25%")
        self.assertEqual(len(report.sections), 4)
        self.assertEqual(
            tuple(section.title for section in report.sections),
            (
                "أكثر 10 موظفين استئذانًا",
                "أكثر 10 موظفين في مهمات العمل",
                "أكثر 10 موظفين غيابًا",
                "أكثر 10 موظفين انضباطًا",
            ),
        )
        self.assertEqual(report.sections[3].rows[0][-1], "25%")
        self.assertContains(response, "أكثر 10 موظفين انضباطًا")

        params["output_format"] = "xlsx"
        exported = self.client.get(reverse("attendance:report_builder"), params)
        workbook = load_workbook(BytesIO(exported.content), read_only=True)
        self.assertEqual(len(workbook.sheetnames), 5)
        self.assertIn("أكثر 10 موظفين غيابًا", workbook.sheetnames)
        workbook.close()

    def test_general_manager_without_department_scopes_can_view_and_export_reports(self):
        self.employee.employment_assignments.all().delete()
        record = self._record(source_status="غياب")
        calculate_records(
            records=[record], requested_by=self.user, import_batch=self.batch
        )
        manager = get_user_model().objects.create_user(
            username="reports-general-manager",
            password="Strong-Test-Pass-2026",
        )
        UserRole.objects.create(
            user=manager,
            role=Role.objects.get(code="general_manager"),
        )
        self.assertFalse(manager.department_scopes.exists())
        self.client.force_login(manager)
        params = {
            "report_type": "comprehensive",
            "limit": "10",
            "output_format": "preview",
        }

        preview = self.client.get(reverse("attendance:report_builder"), params)

        self.assertEqual(preview.status_code, 200)
        self.assertContains(preview, "جميع الموظفين")
        self.assertEqual(dict(preview.context["report"].summary)["عدد الموظفين"], 1)
        self.assertEqual(dict(preview.context["report"].summary)["أيام الغياب"], 1)
        params["output_format"] = "xlsx"
        exported = self.client.get(reverse("attendance:report_builder"), params)
        self.assertEqual(exported.status_code, 200)
        self.assertEqual(
            exported["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_weekly_holidays_are_hidden_from_dashboard_and_reports(self):
        records = [
            self._record(
                attendance_date=date(2026, 6, 28),
                row_number=1,
                source_status="عطلة الأسبوع",
            ),
            self._record(
                attendance_date=date(2026, 6, 29),
                row_number=2,
                source_status="غياب",
            ),
        ]
        calculate_records(
            records=records, requested_by=self.user, import_batch=self.batch
        )
        self.client.force_login(self.user)

        overview = self.client.get(reverse("attendance:report_overview"))
        employee_report = self.client.get(
            reverse("attendance:employee_report"),
            {
                "employee": self.employee.id,
                "date_from": "2026-06-28",
                "date_to": "2026-06-29",
            },
        )
        dashboard = self.client.get(reverse("core:dashboard"))

        self.assertEqual(overview.context["summary"]["records"], 1)
        self.assertEqual(dict(employee_report.context["summary"])["إجمالي الأيام"], 1)
        self.assertNotContains(employee_report, "عطلة الأسبوع")
        unit = dashboard.context["unit_statuses"][0]
        self.assertEqual(unit["value"], 0)

    def test_employee_report_rejects_employee_outside_user_scope(self):
        record = self._record()
        calculate_records(
            records=[record], requested_by=self.user, import_batch=self.batch
        )
        scoped_user = get_user_model().objects.create_user(
            username="employee-report-viewer",
            password="Strong-Test-Pass-2026",
        )
        UserDepartmentScope.objects.create(
            user=scoped_user,
            department=self.department,
            access_level=UserDepartmentScope.AccessLevel.VIEW,
            valid_from=timezone.now() - timedelta(days=1),
        )
        outside_employee = Employee.objects.create(full_name_ar="موظف خارج النطاق")
        self.client.force_login(scoped_user)

        response = self.client.get(
            reverse("attendance:employee_report"),
            {
                "employee": outside_employee.id,
                "date_from": "2026-06-28",
                "date_to": "2026-07-02",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertIsNone(response.context["summary"])
        self.assertIn("employee", response.context["form"].errors)
        self.assertNotContains(response, outside_employee.full_name_ar)

    def test_employee_login_submit_and_department_head_approval(self):
        absence = self._record(
            source_status="غياب",
            in_location="بوابة الحضور",
            out_location="بوابة الانصراف",
        )
        calculate_records(records=[absence], requested_by=self.user, import_batch=self.batch)
        clarification = ClarificationRequest.objects.get()

        login_response = self.client.post(
            reverse("accounts:employee_login"),
            {"national_id": self.national_id},
        )
        self.assertRedirects(login_response, reverse("violations:employee_portal"))
        self.employee.refresh_from_db()
        self.assertIsNotNone(self.employee.user_id)
        portal = self.client.get(reverse("violations:employee_portal"))
        self.assertContains(portal, "موظف الاختبار")
        self.assertContains(portal, "طلبات الإفادة")
        self.assertContains(portal, "مكان الحضور")
        self.assertContains(portal, "بوابة الحضور")
        self.assertContains(portal, "مكان الانصراف")
        self.assertContains(portal, "بوابة الانصراف")
        protected_page = self.client.get(reverse("accounts:user_list"))
        self.assertRedirects(protected_page, reverse("violations:employee_portal"))

        submit_response = self.client.post(
            reverse("violations:employee_clarification", args=(clarification.id,)),
            {"explanation": "كنت في إجازة مرضية مثبتة"},
        )
        self.assertRedirects(submit_response, reverse("violations:employee_portal"))
        clarification.refresh_from_db()
        self.assertEqual(clarification.status, ClarificationRequest.Status.AWAITING_MANAGER)

        mission = self._record(
            source_status="مهمة عمل رسمية",
            attendance_date=date(2026, 6, 29),
            row_number=2,
        )
        calculate_records(
            records=[mission], requested_by=self.user, import_batch=self.batch
        )

        head_user = get_user_model().objects.create_user(
            username="clarification-head", password="Strong-Test-Pass-2026"
        )
        head_employee = Employee.objects.create(
            full_name_ar="رئيس قسم الإفادات", user=head_user
        )
        self.department.department_head = head_employee
        self.department.save(update_fields=("department_head", "updated_at"))
        UserRole.objects.create(
            user=head_user,
            role=Role.objects.get(code="department_head"),
            valid_from=timezone.now(),
        )
        self.client.force_login(head_user)

        manager_page = self.client.get(reverse("violations:manager_dashboard"))
        self.assertContains(manager_page, "موظف الاختبار")
        head_dashboard = self.client.get(reverse("core:dashboard"))
        head_stats = {
            item["title"]: item for item in head_dashboard.context["dashboard_stats"]
        }
        self.assertEqual(head_stats["مهمات العمل"]["value"], 1)
        self.assertContains(head_dashboard, reverse("violations:work_mission_list"))
        manager_missions = self.client.get(reverse("violations:work_mission_list"))
        self.assertEqual(manager_missions.status_code, 200)
        self.assertEqual(manager_missions.context["mission_total"], 1)
        self.assertContains(manager_missions, "موظف الاختبار")
        review_response = self.client.post(
            reverse("violations:manager_review", args=(clarification.id,)),
            {"decision": "approve", "comment": "تمت مراجعة الشاهد"},
        )
        self.assertRedirects(review_response, reverse("violations:manager_dashboard"))
        clarification.refresh_from_db()
        self.assertEqual(clarification.status, ClarificationRequest.Status.APPROVED)

        general_manager = get_user_model().objects.create_user(
            username="general-manager", password="Strong-Test-Pass-2026"
        )
        UserRole.objects.create(
            user=general_manager,
            role=Role.objects.get(code="general_manager"),
            valid_from=timezone.now(),
        )
        self.client.force_login(general_manager)
        executive_page = self.client.get(reverse("violations:executive_dashboard"))
        self.assertEqual(executive_page.status_code, 200)
        self.assertContains(executive_page, self.department.name_ar)
        self.assertContains(executive_page, "المعتمدة")
        general_dashboard = self.client.get(reverse("core:dashboard"))
        general_stats = {
            item["title"]: item for item in general_dashboard.context["dashboard_stats"]
        }
        self.assertEqual(general_stats["مهمات العمل"]["value"], 1)
        self.assertContains(general_dashboard, reverse("violations:work_mission_list"))
        executive_missions = self.client.get(reverse("violations:work_mission_list"))
        self.assertEqual(executive_missions.status_code, 200)
        self.assertContains(executive_missions, "موظف الاختبار")

    def test_department_head_sees_current_assignment_clarifications_without_snapshot(self):
        absence = self._record(source_status="غياب")
        calculate_records(
            records=[absence], requested_by=self.user, import_batch=self.batch
        )
        clarification = ClarificationRequest.objects.get()
        DailyAttendanceResult.objects.update(department=None)
        ClarificationRequest.objects.update(department=None)
        head_user = get_user_model().objects.create_user(
            username="assignment-based-head",
            password="Strong-Test-Pass-2026",
        )
        head_employee = Employee.objects.create(
            full_name_ar="رئيس القسم حسب الإسناد",
            user=head_user,
        )
        EmploymentAssignment.objects.create(
            employee=head_employee,
            department=self.department,
            valid_from=date(2026, 1, 1),
            is_primary=True,
        )
        UserRole.objects.create(
            user=head_user,
            role=Role.objects.get(code="department_head"),
            valid_from=timezone.now(),
        )
        self.client.force_login(head_user)

        dashboard = self.client.get(reverse("violations:manager_dashboard"))
        review = self.client.get(
            reverse("violations:manager_review", args=(clarification.id,))
        )

        self.assertEqual(dashboard.status_code, 200)
        self.assertEqual(dashboard.context["summary"]["total"], 1)
        self.assertContains(dashboard, self.employee.full_name_ar)
        self.assertEqual(review.status_code, 200)

    def test_reviewing_user_delete_is_blocked_with_safe_message(self):
        absence = self._record(source_status="غياب")
        calculate_records(records=[absence], requested_by=self.user, import_batch=self.batch)
        clarification = ClarificationRequest.objects.get()
        reviewer = get_user_model().objects.create_user(
            username="protected-reviewer", password="Strong-Test-Pass-2026"
        )
        clarification.status = ClarificationRequest.Status.APPROVED
        clarification.reviewed_by = reviewer
        clarification.reviewed_at = timezone.now()
        clarification.save(
            update_fields=("status", "reviewed_by", "reviewed_at", "updated_at")
        )
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("accounts:user_delete", args=(reviewer.id,)), follow=True
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "مسجل كمعتمد لإفادات سابقة")
        self.assertTrue(get_user_model().objects.filter(pk=reviewer.id).exists())

    def test_employee_portal_counts_work_missions_from_sheet_status(self):
        first = self._record(source_status="مهمة عمل رسمية")
        second = self._record(
            source_status="مهمه عمل خارجية",
            attendance_date=date(2026, 6, 29),
            row_number=2,
        )
        calculate_records(
            records=[first, second], requested_by=self.user, import_batch=self.batch
        )
        employee_user = get_user_model().objects.create_user(
            username="work-mission-employee", password="Strong-Test-Pass-2026"
        )
        self.employee.user = employee_user
        self.employee.save(update_fields=("user", "updated_at"))
        self.client.force_login(employee_user)

        response = self.client.get(reverse("violations:employee_portal"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["stats"]["work_missions"], 2)
        self.assertContains(response, "مهمة عمل")

    def test_recalculation_supersedes_previous_result(self):
        record = self._record()
        calculate_records(records=[record], requested_by=self.user, import_batch=self.batch)
        calculate_records(records=[record], requested_by=self.user, import_batch=self.batch)

        self.assertEqual(CalculationRun.objects.count(), 2)
        self.assertEqual(DailyAttendanceResult.objects.count(), 2)
        self.assertEqual(DailyAttendanceResult.objects.filter(is_current=True).count(), 1)
        self.assertEqual(DailyAttendanceResult.objects.get(is_current=True).version, 2)

    def test_primary_location_change_does_not_change_location_comparison_rule(self):
        record = self._record(in_location="فرع آخر", out_location="فرع آخر")
        calculate_records(records=[record], requested_by=self.user, import_batch=self.batch)
        new_location = Location.objects.create(
            code="L-2",
            name_ar="فرع آخر",
            location_type=Location.LocationType.BRANCH,
        )
        self.client.force_login(self.user)

        response = self.client.post(
            reverse("organization:employee_edit", args=(self.employee.id,)),
            {
                "full_name_ar": self.employee.full_name_ar,
                "employee_number": "",
                "mobile": "",
                "department": str(self.department.id),
                "location": str(new_location.id),
                "location_effective_date": record.attendance_date.isoformat(),
                "manager_employee": "",
                "employment_status": self.employee.employment_status,
            },
        )

        self.assertEqual(response.status_code, 302)
        current = DailyAttendanceResult.objects.get(is_current=True)
        self.assertEqual(current.version, 1)
        self.assertIsNone(current.primary_location)
        self.assertEqual(current.location_status, DailyAttendanceResult.LocationStatus.MATCHED)
        self.assertEqual(DailyAttendanceResult.objects.filter(is_current=False).count(), 0)

    def test_reports_search_by_name_or_national_id_and_limit_full_display_to_admin(self):
        record = self._record(in_location="فرع آخر")
        calculate_records(records=[record], requested_by=self.user, import_batch=self.batch)
        self.client.force_login(self.user)

        for url_name in ("attendance:record_list", "attendance:outside_location_report"):
            with self.subTest(url_name=url_name):
                by_national_id = self.client.post(
                    reverse(url_name), {"employee_search": self.national_id}
                )
                self.assertEqual(by_national_id.status_code, 200)
                self.assertContains(by_national_id, self.employee.full_name_ar)
                self.assertContains(by_national_id, self.national_id)

                by_name = self.client.post(
                    reverse(url_name), {"employee_search": "موظف الاختبار"}
                )
                self.assertContains(by_name, self.employee.full_name_ar)

        scoped_user = get_user_model().objects.create_user(
            username="report-viewer", password="Strong-Test-Pass-2026"
        )
        UserDepartmentScope.objects.create(
            user=scoped_user,
            department=self.department,
            access_level=UserDepartmentScope.AccessLevel.VIEW,
            valid_from=timezone.now() - timedelta(days=1),
        )
        self.client.force_login(scoped_user)

        masked_response = self.client.post(
            reverse("attendance:outside_location_report"),
            {"employee_search": self.national_id},
        )

        self.assertEqual(masked_response.status_code, 200)
        self.assertContains(masked_response, "******6789")
        self.assertNotContains(masked_response, self.national_id)

    def test_daily_record_uses_current_department_when_historical_snapshot_is_missing(self):
        assignment = self.employee.employment_assignments.get(is_primary=True)
        assignment.valid_from = date(2026, 7, 12)
        assignment.save(update_fields=("valid_from", "updated_at"))
        record = self._record()
        calculate_records(records=[record], requested_by=self.user, import_batch=self.batch)
        self.assertIsNone(DailyAttendanceResult.objects.get().department_id)

        scoped_user = get_user_model().objects.create_user(
            username="department-report-viewer",
            password="Strong-Test-Pass-2026",
        )
        UserDepartmentScope.objects.create(
            user=scoped_user,
            department=self.department,
            access_level=UserDepartmentScope.AccessLevel.VIEW,
            valid_from=timezone.now() - timedelta(days=1),
        )
        self.client.force_login(scoped_user)

        response = self.client.post(
            reverse("attendance:record_list"),
            {"department": str(self.department.id)},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.employee.full_name_ar)
        self.assertContains(response, self.department.name_ar)
