from __future__ import annotations

from datetime import date, time, timedelta
from io import BytesIO

from django.test import SimpleTestCase
from openpyxl import Workbook

from attendance.services.weekly_report_parser import parse_weekly_report


HEADERS = (
    "السجل المدني",
    "الاسم",
    "المسمى الوظيفي",
    "التاريخ",
    "حالة التحضير",
    "ساعات الدوام",
    "وقت الحضور",
    "مكان الحضور",
    "توقيت الانصراف",
    "مكان الانصراف",
    "ساعات الدوام الفعلي",
    "انصراف مبكر",
    "النقص في الدوام",
    "حضور مبكر",
)


def _daily_row(
    *,
    national_id: object = "1023456789",
    employee_name: object = "موظف تجريبي",
    job_title: object = "محلل",
    attendance_date: object = "2026/07/05",
    source_status: object = "حاضر",
    scheduled_duration: object = "08:00",
    check_in: object = "08:00",
    check_in_location: object = "المقر الرئيسي",
    check_out: object = "16:00",
    check_out_location: object = "المقر الرئيسي",
    actual_work_duration: object = "08:00",
    early_departure_duration: object = "00:00",
    shortfall_duration: object = "00:00",
    early_arrival_duration: object = "00:00",
) -> list[object]:
    return [
        national_id,
        employee_name,
        job_title,
        attendance_date,
        source_status,
        scheduled_duration,
        check_in,
        check_in_location,
        check_out,
        check_out_location,
        actual_work_duration,
        early_departure_duration,
        shortfall_duration,
        early_arrival_duration,
    ]


def _workbook_bytes(
    *rows: list[object],
    merge_employee_rows: bool = False,
) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "تقرير الحضور الأسبوعي"
    worksheet.append(
        ["تقرير الفترة 2026/07/05 - 2026/07/09", *([None] * 13)]
    )
    worksheet.append([None] * len(HEADERS))
    worksheet.append(HEADERS)
    for row in rows:
        worksheet.append(row)

    if merge_employee_rows:
        for column in range(1, 4):
            worksheet.merge_cells(
                start_row=4,
                start_column=column,
                end_row=5,
                end_column=column,
            )

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


class WeeklyReportParserTests(SimpleTestCase):
    def test_month_period_is_read_from_title_above_sheet(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["تقرير الحضور من 2026/07/01 إلى 2026/07/31"])
        worksheet.append(HEADERS)
        worksheet.append(_daily_row(attendance_date="2026/07/15"))
        output = BytesIO()
        workbook.save(output)
        workbook.close()

        report = parse_weekly_report(output.getvalue())

        self.assertEqual(report.period_start, date(2026, 7, 1))
        self.assertEqual(report.period_end, date(2026, 7, 31))
        self.assertIn("2026/07/01", report.period_title)

    def test_period_is_not_inferred_from_daily_rows_when_title_has_no_dates(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["تقرير حضور شهر يوليو"])
        worksheet.append(HEADERS)
        worksheet.append(_daily_row(attendance_date="2026/07/15"))
        output = BytesIO()
        workbook.save(output)
        workbook.close()

        report = parse_weekly_report(output.getvalue())

        self.assertIsNone(report.period_start)
        self.assertIsNone(report.period_end)
        self.assertIn("period_not_found", {issue.code for issue in report.issues})

    def test_one_employee_with_multiple_daily_rows(self):
        report = parse_weekly_report(
            _workbook_bytes(
                _daily_row(attendance_date="2026/07/05"),
                _daily_row(
                    national_id=None,
                    employee_name=None,
                    job_title=None,
                    attendance_date="2026/07/06",
                    check_in="08:15",
                    check_out="16:10",
                ),
            )
        )

        self.assertEqual(report.period_start, date(2026, 7, 5))
        self.assertEqual(report.period_end, date(2026, 7, 9))
        self.assertEqual(report.employee_count, 1)
        self.assertEqual(len(report.rows), 2)
        self.assertEqual(report.rows[1].national_id, "1023456789")
        self.assertEqual(report.rows[1].employee_name, "موظف تجريبي")
        self.assertEqual(report.rows[1].job_title, "محلل")
        self.assertEqual(report.rows[1].attendance_date, date(2026, 7, 6))

    def test_multiple_employees_create_distinct_contexts(self):
        report = parse_weekly_report(
            _workbook_bytes(
                _daily_row(national_id="1023456789", employee_name="الموظف الأول"),
                _daily_row(
                    national_id="1098765432",
                    employee_name="الموظف الثاني",
                    attendance_date="2026/07/06",
                ),
            )
        )

        self.assertEqual(report.employee_count, 2)
        self.assertEqual(
            [row.national_id for row in report.rows],
            ["1023456789", "1098765432"],
        )
        self.assertEqual(
            [row.employee_name for row in report.rows],
            ["الموظف الأول", "الموظف الثاني"],
        )

    def test_summary_rows_are_skipped_and_counted(self):
        report = parse_weekly_report(
            _workbook_bytes(
                _daily_row(),
                ["المجموع", *([None] * 13)],
            )
        )

        self.assertEqual(len(report.rows), 1)
        self.assertEqual(report.summary_row_count, 1)

    def test_blank_rows_are_skipped_and_counted(self):
        report = parse_weekly_report(
            _workbook_bytes(
                _daily_row(attendance_date="2026/07/05"),
                [None] * len(HEADERS),
                _daily_row(
                    national_id=None,
                    employee_name=None,
                    job_title=None,
                    attendance_date="2026/07/06",
                ),
            )
        )

        self.assertEqual(len(report.rows), 2)
        self.assertEqual(report.ignored_row_count, 4)

    def test_merged_employee_cells_are_carried_forward(self):
        report = parse_weekly_report(
            _workbook_bytes(
                _daily_row(attendance_date="2026/07/05"),
                _daily_row(
                    national_id=None,
                    employee_name=None,
                    job_title=None,
                    attendance_date="2026/07/06",
                ),
                merge_employee_rows=True,
            )
        )

        self.assertEqual(len(report.rows), 2)
        self.assertEqual(report.rows[1].national_id, report.rows[0].national_id)
        self.assertEqual(report.rows[1].employee_name, report.rows[0].employee_name)
        self.assertEqual(report.rows[1].job_title, report.rows[0].job_title)

    def test_arabic_and_persian_digits_and_null_markers_are_normalized(self):
        report = parse_weekly_report(
            _workbook_bytes(
                _daily_row(
                    national_id="١٠٢٣٤٥٦٧٨٩",
                    attendance_date="۲۰۲۶/۰۷/۰۵",
                    source_status="-",
                    scheduled_duration="٠٨:٠٠",
                    check_in="٠٨:٣٠",
                    check_in_location="—",
                    check_out="۱۶:۱۵",
                    check_out_location="-",
                    actual_work_duration="۰۷:۴۵",
                    early_departure_duration="–",
                    shortfall_duration="لا يوجد",
                    early_arrival_duration="-",
                )
            )
        )

        row = report.rows[0]
        self.assertEqual(row.national_id, "1023456789")
        self.assertEqual(row.attendance_date, date(2026, 7, 5))
        self.assertEqual(row.check_in, time(8, 30))
        self.assertEqual(row.check_out, time(16, 15))
        self.assertEqual(row.scheduled_duration, timedelta(hours=8))
        self.assertEqual(row.actual_work_duration, timedelta(hours=7, minutes=45))
        self.assertIsNone(row.source_status)
        self.assertIsNone(row.check_in_location)
        self.assertIsNone(row.check_out_location)
        self.assertIsNone(row.early_departure_duration)
        self.assertIsNone(row.shortfall_duration)
        self.assertIsNone(row.early_arrival_duration)

    def test_invalid_date_and_times_are_reported(self):
        report = parse_weekly_report(
            _workbook_bytes(
                _daily_row(
                    attendance_date="2026/99/99",
                    check_in="25:00",
                    check_out="وقت غير صالح",
                )
            )
        )

        row = report.rows[0]
        self.assertIsNone(row.attendance_date)
        self.assertIsNone(row.check_in)
        self.assertIsNone(row.check_out)
        self.assertEqual(
            {issue.code for issue in row.issues},
            {"invalid_date", "invalid_check_in", "invalid_check_out"},
        )

    def test_incomplete_daily_row_is_reported(self):
        report = parse_weekly_report(
            _workbook_bytes(
                _daily_row(
                    source_status=None,
                    scheduled_duration=None,
                    check_in=None,
                    check_in_location=None,
                    check_out=None,
                    check_out_location=None,
                    actual_work_duration=None,
                    early_departure_duration=None,
                    shortfall_duration=None,
                    early_arrival_duration=None,
                )
            )
        )

        row = report.rows[0]
        self.assertEqual(row.attendance_date, date(2026, 7, 5))
        self.assertIn("incomplete_daily_row", {issue.code for issue in row.issues})
    def test_official_report_header_spelling_variants_are_supported(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "WorkSheet"
        worksheet.append([None] * 14)
        worksheet.append([None] * 6 + ["الحضور والانصراف لموظفين من: 2026/06/28 الى 2026/07/02"] + [None] * 7)
        worksheet.append([None] * 14)
        worksheet.append((
            "السجل المدني",
            "الإسم",
            "المسمى الوظيفي",
            "التاريخ",
            "حالة التحضير",
            "ساعات الدوام",
            "توقيت الحضور",
            "مكان الحضور",
            "توقيت الإنصراف",
            "مكان الانصراف",
            "ساعات الدوام الفعلي",
            "انصراف مبكر",
            "النقص في الدوام",
            "حضور مبكر",
        ))
        worksheet.append(_daily_row(attendance_date="الأحد 28-06-26"))
        output = BytesIO()
        workbook.save(output)
        workbook.close()

        report = parse_weekly_report(output.getvalue())

        self.assertEqual(report.period_start, date(2026, 6, 28))
        self.assertEqual(report.period_end, date(2026, 7, 2))
        self.assertEqual(len(report.rows), 1)
        self.assertEqual(report.rows[0].attendance_date, date(2026, 6, 28))
        self.assertEqual(report.rows[0].employee_name, "موظف تجريبي")
